"""Tests for src/imdr/domains/econ/seek_jobads.py parse + transform layer.

No network calls — synthetic openpyxl workbooks built in-memory mirror the
real SEEK workbook column layouts (verified 2026-07-14, see the module
docstring), and a synthetic HTML snippet mirrors the report page's download
links.

Covered:
- discover_download_urls pulls both graphassets.com links off the report
  page, and raises a clear error when either is missing.
- _industry_suffix slugifies SEEK's classification labels into IND_ codes.
- parse_job_ad_sheet / parse_salary_sheet extract raw rows, filtering to
  country=Australia.
- job_ad_rows_to_records / salary_rows_to_records build (indicators,
  observations), respecting since/until windows and skipping null values,
  and correctly split state-cut vs industry-cut vs national in the salary
  table's shared flat layout.
"""

from __future__ import annotations

import datetime

import openpyxl
import pytest

from imdr.domains.econ.seek_jobads import (
    _industry_suffix,
    _job_ad_suffix,
    _salary_suffix,
    discover_download_urls,
    job_ad_rows_to_records,
    parse_job_ad_sheet,
    parse_salary_sheet,
    salary_rows_to_records,
)


def _report_page_html(*, employment_url: str | None, salary_url: str | None) -> str:
    emp_link = (
        f'<a href="{employment_url}" class="btn">Download the latest SEEK Employment data here</a>'
        if employment_url else ""
    )
    sal_link = (
        f'<a href="{salary_url}" class="btn">Download the latest SEEK Advertised Salary data here</a>'
        if salary_url else ""
    )
    return f"<html><body>{emp_link}{sal_link}</body></html>"


class TestDiscoverDownloadUrls:
    def test_extracts_both_links(self) -> None:
        html = _report_page_html(
            employment_url="https://ap-southeast-2-seek-apac.graphassets.com/AEz/emp123",
            salary_url="https://ap-southeast-2-seek-apac.graphassets.com/AEz/sal456",
        )
        emp_url, sal_url = discover_download_urls(html)
        assert emp_url == "https://ap-southeast-2-seek-apac.graphassets.com/AEz/emp123"
        assert sal_url == "https://ap-southeast-2-seek-apac.graphassets.com/AEz/sal456"

    def test_raises_when_employment_link_missing(self) -> None:
        html = _report_page_html(
            employment_url=None,
            salary_url="https://ap-southeast-2-seek-apac.graphassets.com/AEz/sal456",
        )
        with pytest.raises(RuntimeError, match="SEEK Employment data"):
            discover_download_urls(html)

    def test_raises_when_salary_link_missing(self) -> None:
        html = _report_page_html(
            employment_url="https://ap-southeast-2-seek-apac.graphassets.com/AEz/emp123",
            salary_url=None,
        )
        with pytest.raises(RuntimeError, match="SEEK Advertised Salary data"):
            discover_download_urls(html)


class TestIndustrySuffix:
    def test_slugifies_ampersand_and_comma(self) -> None:
        assert _industry_suffix("Information & Communication Technology") == (
            "IND_INFORMATION_AND_COMMUNICATION_TECHNOLOGY"
        )
        assert _industry_suffix("Advertising, Arts & Media") == "IND_ADVERTISING_ARTS_AND_MEDIA"
        assert _industry_suffix("Accounting") == "IND_ACCOUNTING"


class TestJobAdSuffix:
    def test_total_maps_to_national(self) -> None:
        assert _job_ad_suffix("Total") == ("NATIONAL", "National")

    def test_known_state_maps_to_state_suffix(self) -> None:
        assert _job_ad_suffix("NSW") == ("STATE_NSW", "NSW")

    def test_unknown_state_returns_none(self) -> None:
        assert _job_ad_suffix("Wollongong") is None


class TestSalarySuffix:
    def test_national_headline(self) -> None:
        assert _salary_suffix("Total", "Total") == ("NATIONAL", "National")

    def test_state_cut(self) -> None:
        assert _salary_suffix("VIC", "Total") == ("STATE_VIC", "VIC")

    def test_industry_cut(self) -> None:
        assert _salary_suffix("Total", "Legal") == ("IND_LEGAL", "Legal")

    def test_unexpected_cross_combo_returns_none(self) -> None:
        assert _salary_suffix("VIC", "Legal") is None


def _job_ad_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SEEK Job Ad Index")
    ws.append([
        "DATE", "COUNTRY", "STATE", "ADS_SA_INDEX", "ADS_TREND_INDEX",
        "ADS_SA_GROWTH_MONTH", "ADS_SA_GROWTH_PCP",
        "ADS_TREND_GROWTH_MONTH", "ADS_TREND_GROWTH_PCP",
    ])
    ws.append([datetime.datetime(2026, 4, 1), "Australia", "Total", 110.0, 111.0, 0.01, 0.02, 0.01, 0.02])
    ws.append([datetime.datetime(2026, 5, 1), "Australia", "Total", 111.6, 112.8, -0.02, -0.06, -0.01, -0.04])
    ws.append([datetime.datetime(2026, 5, 1), "Australia", "NSW", 105.0, 106.0, 0.0, 0.0, 0.0, 0.0])
    ws.append([datetime.datetime(2026, 5, 1), "New Zealand", "Total", 90.0, 91.0, 0.0, 0.0, 0.0, 0.0])
    return wb


def _salary_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append([
        "date", "country", "state", "classification", "salary_sa_index",
        "salary_sa_growth_month", "salary_sa_growth_pcp",
        "salary_trend_index", "salary_trend_growth_month", "salary_trend_growth_pcp",
    ])
    ws.append([datetime.datetime(2026, 5, 1), "Australia", "Total", "Total", 100.0, 0.001, 0.02, 100.5, 0.001, 0.02])
    ws.append([datetime.datetime(2026, 5, 1), "Australia", "VIC", "Total", 98.0, 0.001, 0.02, 98.5, 0.001, 0.02])
    ws.append([datetime.datetime(2026, 5, 1), "Australia", "Total", "Legal", 120.0, 0.001, 0.02, None, None, None])
    return wb


class TestParseJobAdSheet:
    def test_filters_to_australia_and_extracts_fields(self) -> None:
        rows = parse_job_ad_sheet(_job_ad_workbook())
        assert len(rows) == 3
        assert all(r["date"].year == 2026 for r in rows)
        nsw = next(r for r in rows if r["state"] == "NSW")
        assert nsw["ads_sa_index"] == 105.0
        assert nsw["ads_trend_index"] == 106.0


class TestParseSalarySheet:
    def test_filters_to_australia_and_extracts_fields(self) -> None:
        rows = parse_salary_sheet(_salary_workbook())
        assert len(rows) == 3
        legal = next(r for r in rows if r["classification"] == "Legal")
        assert legal["state"] == "Total"
        assert legal["salary_sa_index"] == 120.0
        assert legal["salary_trend_index"] is None


class TestJobAdRowsToRecords:
    def test_builds_national_and_state_series_sa_and_trend(self) -> None:
        raw = parse_job_ad_sheet(_job_ad_workbook())
        indicators, observations = job_ad_rows_to_records(raw)
        codes = {i.imdr_code for i in indicators}
        assert codes == {
            "SEEK.JOBADS.INDEX.NATIONAL.AU",
            "SEEK.JOBADS.INDEX_TREND.NATIONAL.AU",
            "SEEK.JOBADS.INDEX.STATE_NSW.AU",
            "SEEK.JOBADS.INDEX_TREND.STATE_NSW.AU",
        }
        national_sa = [
            o for o in observations if o.imdr_code == "SEEK.JOBADS.INDEX.NATIONAL.AU"
        ]
        assert len(national_sa) == 2
        latest = next(o for o in national_sa if o.obs_date == datetime.date(2026, 5, 1))
        assert latest.value == 111.6

    def test_indicator_metadata(self) -> None:
        raw = parse_job_ad_sheet(_job_ad_workbook())
        indicators, _ = job_ad_rows_to_records(raw)
        sa = next(i for i in indicators if i.imdr_code == "SEEK.JOBADS.INDEX.NATIONAL.AU")
        trend = next(i for i in indicators if i.imdr_code == "SEEK.JOBADS.INDEX_TREND.NATIONAL.AU")
        assert sa.category == "labour"
        assert sa.frequency == "MONTHLY"
        assert sa.country_iso == "AU"
        assert sa.unit == "index"
        assert sa.is_seasonally_adjusted is True
        assert trend.is_seasonally_adjusted is False

    def test_since_until_window_filters_observations(self) -> None:
        raw = parse_job_ad_sheet(_job_ad_workbook())
        _, observations = job_ad_rows_to_records(
            raw, since=datetime.date(2026, 5, 1), until=datetime.date(2026, 5, 31),
        )
        obs_dates = {o.obs_date for o in observations}
        assert obs_dates == {datetime.date(2026, 5, 1)}


class TestSalaryRowsToRecords:
    def test_builds_national_state_and_industry_series(self) -> None:
        raw = parse_salary_sheet(_salary_workbook())
        indicators, observations = salary_rows_to_records(raw)
        codes = {i.imdr_code for i in indicators}
        assert codes == {
            "SEEK.SALARY.INDEX.NATIONAL.AU",
            "SEEK.SALARY.INDEX_TREND.NATIONAL.AU",
            "SEEK.SALARY.INDEX.STATE_VIC.AU",
            "SEEK.SALARY.INDEX_TREND.STATE_VIC.AU",
            "SEEK.SALARY.INDEX.IND_LEGAL.AU",
        }
        # Legal's trend value is None in the fixture -- no observation emitted for it.
        legal_trend = [o for o in observations if o.imdr_code == "SEEK.SALARY.INDEX_TREND.IND_LEGAL.AU"]
        assert legal_trend == []

    def test_indicator_metadata(self) -> None:
        raw = parse_salary_sheet(_salary_workbook())
        indicators, _ = salary_rows_to_records(raw)
        ind = next(i for i in indicators if i.imdr_code == "SEEK.SALARY.INDEX.IND_LEGAL.AU")
        assert ind.category == "labour"
        assert ind.frequency == "MONTHLY"
        assert ind.country_iso == "AU"
        assert "Legal" in ind.display_name
