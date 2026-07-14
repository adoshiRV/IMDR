"""APRA Monthly ADI Statistics (MADIS) — housing loan book by ADI.

Source: https://www.apra.gov.au/monthly-authorised-deposit-taking-institution-statistics

APRA publishes two files off that page each month:
  - the current-month workbook (wide "Table 2" layout, one sheet per month) — not used here.
  - the **back-series** workbook: a single long-format sheet ("Table 1") with
    one row per (Period, ABN, Institution Name), covering the full history
    (verified 2026-07-14: 87 monthly periods, 2019-03-31 -> 2026-05-31,
    10,886 rows, 195 ADIs). This already carries the latest month, so it is
    the sole source we need -- no need to also parse the current-month file.

The page also still hosts a legacy pre-MADIS-rename file
("Monthly banking statistics June 2019 back series", published 2019-09-08)
under a different anchor-text prefix ("Monthly banking statistics" vs
"Monthly authorised deposit-taking institution statistics") -- excluded
by `discover_backseries_url`'s text match.

Confirmed public / plain HTTP GET, no auth, no browser needed (curl 200 OK
verified 2026-07-14) -- same as the quarterly ADI performance stats in
scripts/econ/au/govt/fetch_apra_quarterly.py, not subject to the AOFM
corp-firewall block.

Columns of interest in Table 1 (($million), header row 2, data from row 3):
  Period                                              -- month-end date
  Institution Name                                    -- ADI legal name
  Loans to households: Housing: Owner-occupied        -- AUD million
  Loans to households: Housing: Investment            -- AUD million

Big-4 institution names as they appear in the file (verified 2026-07-14,
stable ABN per name across the full back series):
  Australia and New Zealand Banking Group Limited  -> ANZ
  Commonwealth Bank of Australia                   -> CBA
  National Australia Bank Limited                  -> NAB
  Westpac Banking Corporation                      -> WBC

No system-total row exists in the back-series long format (unlike the
current-month wide file's "TOTAL" row in Table 2) -- a system aggregate
would have to be computed by summing all ~195 ADIs, which is out of scope
here; only the big-4 (plus any BANK_MAP additions) are emitted.
"""
from __future__ import annotations

import datetime
import math
import re
from pathlib import Path

from imdr.domains.econ.schema import IndicatorRow, ObservationRow

UTC = datetime.timezone.utc

MADIS_PAGE_URL = "https://www.apra.gov.au/monthly-authorised-deposit-taking-institution-statistics"
APRA_BASE = "https://www.apra.gov.au"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) IMDR-apra-madis"

BANK_MAP: dict[str, str] = {
    "Australia and New Zealand Banking Group Limited": "ANZ",
    "Commonwealth Bank of Australia": "CBA",
    "National Australia Bank Limited": "NAB",
    "Westpac Banking Corporation": "WBC",
}

_SHEET_NAME = "Table 1"
_PERIOD_COL = "Period"
_INSTITUTION_COL = "Institution Name"
_OWNER_OCC_COL = "Loans to households: Housing: Owner-occupied"
_INVESTOR_COL = "Loans to households: Housing: Investment"


def discover_backseries_url(html: str) -> str | None:
    """Find the current MADIS back-series XLSX download link on the pub page.

    Matches anchor text starting with the current product name ("Monthly
    authorised deposit-taking institution statistics") AND containing
    "back-series"/"back series" -- excludes both the current-month single
    file (no "back" in its text) and the legacy pre-rename back-series file
    ("Monthly banking statistics ... back series").
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=re.compile(r"\.xlsx", re.I)):
        href = a.get("href", "")
        text = a.get_text(" ", strip=True).lower()
        if not text.startswith("monthly authorised deposit-taking institution statistics"):
            continue
        if "back-series" not in text and "back series" not in text:
            continue
        return href if href.startswith("http") else APRA_BASE + href
    return None


def fetch_page_html(*, timeout: int = 30) -> str:
    import httpx

    resp = httpx.get(MADIS_PAGE_URL, headers={"User-Agent": _UA}, timeout=timeout, follow_redirects=True)
    resp.raise_for_status()
    return resp.text


def fetch_backseries_xlsx(url: str, *, timeout: int = 60) -> bytes:
    import httpx

    resp = httpx.get(url, headers={"User-Agent": _UA}, timeout=timeout, follow_redirects=True)
    resp.raise_for_status()
    return resp.content


def coerce_float(v) -> float | None:
    """Accept any workbook cell; return float or None for blank/non-numeric."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        return None if math.isnan(f) else f
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def coerce_date(v) -> datetime.date | None:
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _find_header_row(rows: list[tuple]) -> int:
    """Return the 0-based index of the header row (the one whose first
    cell reads "Period")."""
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().lower() == _PERIOD_COL.lower():
            return i
    raise RuntimeError(f"could not locate header row (looking for {_PERIOD_COL!r} in col A)")


def parse_backseries_rows(rows: list[tuple]) -> list[dict]:
    """Parse the raw ``Table 1`` rows (as returned by
    ``ws.iter_rows(values_only=True)``) into a flat list of
    ``{period, institution, owner_occ, investor}`` dicts for institutions in
    ``BANK_MAP``.
    """
    header_idx = _find_header_row(rows)
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    try:
        period_col = header.index(_PERIOD_COL)
        institution_col = header.index(_INSTITUTION_COL)
        owner_occ_col = header.index(_OWNER_OCC_COL)
        investor_col = header.index(_INVESTOR_COL)
    except ValueError as exc:
        raise RuntimeError(f"MADIS Table 1 header missing expected column: {exc}") from exc

    out: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or row[institution_col] is None:
            continue
        institution = str(row[institution_col]).strip()
        if institution not in BANK_MAP:
            continue
        period = coerce_date(row[period_col])
        if period is None:
            continue
        out.append({
            "period": period,
            "institution": institution,
            "owner_occ": coerce_float(row[owner_occ_col]),
            "investor": coerce_float(row[investor_col]),
        })
    return out


def parse_backseries_xlsx(path_or_bytes) -> list[dict]:
    import io

    import openpyxl

    src = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    return parse_backseries_rows(rows)


def make_indicator(*, bank_code: str, series: str, display_suffix: str) -> IndicatorRow:
    imdr_code = f"APRA.ADI.{bank_code}.{series}.AU"
    return IndicatorRow(
        imdr_code=imdr_code,
        vendor_name="APRA",
        source_code=f"APRA.MADIS.{bank_code}.{series}",
        display_name=f"APRA MADIS — {bank_code} housing loans: {display_suffix} (AUD million)",
        unit="aud_mn",
        frequency="MONTHLY",
        country_iso="AU",
        category="credit",
        is_seasonally_adjusted=False,
    )


def make_observation(*, imdr_code: str, obs_date: datetime.date, value: float | None) -> ObservationRow:
    now = datetime.datetime.now(UTC)
    return ObservationRow(
        imdr_code=imdr_code, obs_date=obs_date, vintage=0,
        release_date=now, value=value, ingested_at=now,
    )


def build_rows(
    parsed: list[dict],
    *,
    since: datetime.date | None = None,
    until: datetime.date | None = None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """Turn `parse_backseries_xlsx` output into (IndicatorRow, ObservationRow) lists."""
    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for row in parsed:
        d = row["period"]
        if since and d < since:
            continue
        if until and d > until:
            continue
        bank_code = BANK_MAP[row["institution"]]

        for series, display_suffix, value in (
            ("HOUSING_OWNER_OCC", "owner-occupied", row["owner_occ"]),
            ("HOUSING_INVESTOR", "investment", row["investor"]),
        ):
            imdr_code = f"APRA.ADI.{bank_code}.{series}.AU"
            if imdr_code not in seen_codes:
                indicators.append(make_indicator(bank_code=bank_code, series=series, display_suffix=display_suffix))
                seen_codes.add(imdr_code)
            observations.append(make_observation(imdr_code=imdr_code, obs_date=d, value=value))

    return indicators, observations
