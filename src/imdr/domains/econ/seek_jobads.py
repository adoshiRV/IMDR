"""SEEK — Advertised Job Index (job-ad volumes) + Advertised Salary Index.

Source: https://au.seek.com/about/news/article/seek-employment-data -- a
server-rendered React page, so plain ``httpx`` works with no Playwright.
Investigated 2026-07-14: note the hostname -- ``au.seek.com`` (this page) is
reachable with a plain GET, while ``www.seek.com.au`` returns a 403 at the
edge for the same plain request (confirmed not a general network issue:
google.com/abs.gov.au both return 200 from the same client). No login/paywall
on either the report page or the two data downloads below.

The page embeds two "Download the latest ... data here" links pointing at
Hygraph/GraphCMS asset storage (``ap-southeast-2-seek-apac.graphassets.com``).
Each link is a *content-hashed* URL that changes with every monthly release
(confirmed: filenames carry the release date/month), so this fetcher
re-scrapes the report page every run rather than hardcoding a download URL --
see ``discover_download_urls``.

SEEK Job Ad Index workbook (``AU_PUBLISHED_DATASET *.xlsx``):
  Sheet "SEEK Job Ad Index": monthly, 2001-07 -> present (~300 obs/series
  for the national row). Country=Australia only (no NZ in this file). State
  dimension: national Total + 8 states/territories (ACT/NSW/NT/QLD/SA/TAS/
  VIC/WA) -- NO industry breakdown in this workbook. Two variants per cut:
  ADS_SA_INDEX (seasonally adjusted) and ADS_TREND_INDEX (trend). Index base
  = 2016 average = 100 (confirmed from the data: national-Total 2016 average
  is ~99.9; the workbook's own Notes sheet also documents a pending
  re-indexing from a 2013 to a 2016 average base, i.e. 2016=100 is already
  in effect in this download). A third sheet, "SEEK Applications per Ad
  Index" (candidate competition per ad, not ad volume), is present in the
  same file but out of scope for this fetcher.

SEEK Advertised Salary Index workbook (``seek_asi_*_upload.xlsx``):
  Single sheet, monthly, 2015-11 -> present (~127 obs/series). Country=
  Australia only. Two *separate* dimension cuts share one flat table: a
  state cut (``state`` != Total, ``classification`` == Total) and an
  industry/classification cut (``state`` == Total, ``classification`` !=
  Total, 27 industries per SEEK's own classification taxonomy) -- confirmed
  no state x industry cross-tab exists in the published file (0 rows with
  both non-Total). One combined national headline row (state=Total,
  classification=Total). Same SA_INDEX / TREND_INDEX variant split as the
  job-ad workbook.
  WORKAROUND: this workbook has a broken embedded-drawing relationship
  (points at a stripped ``xl/drawings/drawing1.xml`` not present in the zip)
  which raises ``KeyError`` in stock openpyxl on load -- monkeypatched
  around in ``_load_workbook``.

Not built: SEEK Applications-per-Ad Index (candidate competition, not ad
volume/salary); NZ data (SEEK also runs nz.seek.com but this workbook is
AU-only; a distinct NZ dataset would need its own investigation).
"""
from __future__ import annotations

import datetime
import io
import re

import httpx

from imdr.domains.econ.schema import IndicatorRow, ObservationRow

UTC = datetime.timezone.utc
VENDOR_NAME = "SEEK"

REPORT_PAGE_URL = "https://au.seek.com/about/news/article/seek-employment-data"
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0"

JOB_AD_SHEET_NAME = "SEEK Job Ad Index"

STATE_SUFFIX: dict[str, str] = {
    "ACT": "STATE_ACT",
    "NSW": "STATE_NSW",
    "NT": "STATE_NT",
    "QLD": "STATE_QLD",
    "SA": "STATE_SA",
    "TAS": "STATE_TAS",
    "VIC": "STATE_VIC",
    "WA": "STATE_WA",
}

_EMPLOYMENT_LINK_RE = re.compile(
    r'<a href="(https://[^"]+graphassets\.com/[^"]+)"[^>]*>'
    r"Download the latest SEEK Employment data here</a>"
)
_SALARY_LINK_RE = re.compile(
    r'<a href="(https://[^"]+graphassets\.com/[^"]+)"[^>]*>'
    r"Download the latest SEEK Advertised Salary data here</a>"
)


def _industry_suffix(label: str) -> str:
    """"Information & Communication Technology" -> "IND_INFORMATION_AND_COMMUNICATION_TECHNOLOGY"."""
    s = label.upper().replace("&", "AND")
    s = re.sub(r"[^A-Z0-9]+", "_", s).strip("_")
    return f"IND_{s}"


def _make_client() -> httpx.Client:
    return httpx.Client(timeout=30.0, follow_redirects=True, headers={"User-Agent": _UA})


def fetch_report_page_html(client: httpx.Client) -> str:
    r = client.get(REPORT_PAGE_URL)
    r.raise_for_status()
    return r.text


def discover_download_urls(html: str) -> tuple[str, str]:
    """Return (employment_xlsx_url, salary_xlsx_url) scraped off the report page."""
    emp_m = _EMPLOYMENT_LINK_RE.search(html)
    sal_m = _SALARY_LINK_RE.search(html)
    if not emp_m:
        raise RuntimeError(
            "Could not find the 'Download the latest SEEK Employment data here' "
            "link on the report page -- page layout may have changed"
        )
    if not sal_m:
        raise RuntimeError(
            "Could not find the 'Download the latest SEEK Advertised Salary data "
            "here' link on the report page -- page layout may have changed"
        )
    return emp_m.group(1), sal_m.group(1)


def fetch_workbook_bytes(client: httpx.Client, url: str) -> bytes:
    r = client.get(url)
    r.raise_for_status()
    return r.content


def _load_workbook(wb_bytes: bytes):
    import openpyxl
    import openpyxl.reader.excel as _excel_reader

    orig_find_images = _excel_reader.find_images

    def _patched_find_images(archive, path):
        try:
            return orig_find_images(archive, path)
        except KeyError:
            return [], []

    _excel_reader.find_images = _patched_find_images
    try:
        return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    finally:
        _excel_reader.find_images = orig_find_images


def parse_job_ad_sheet(wb) -> list[dict]:
    """Extract raw rows from the "SEEK Job Ad Index" sheet of the employment workbook."""
    ws = wb[JOB_AD_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().upper() for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out: list[dict] = []
    for r in rows[1:]:
        if r[idx["COUNTRY"]] != "Australia":
            continue
        out.append({
            "date": r[idx["DATE"]],
            "state": r[idx["STATE"]],
            "ads_sa_index": r[idx["ADS_SA_INDEX"]],
            "ads_trend_index": r[idx["ADS_TREND_INDEX"]],
        })
    return out


def parse_salary_sheet(wb) -> list[dict]:
    """Extract raw rows from the salary index workbook's single sheet."""
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out: list[dict] = []
    for r in rows[1:]:
        if r[idx["country"]] != "Australia":
            continue
        out.append({
            "date": r[idx["date"]],
            "state": r[idx["state"]],
            "classification": r[idx["classification"]],
            "salary_sa_index": r[idx["salary_sa_index"]],
            "salary_trend_index": r[idx["salary_trend_index"]],
        })
    return out


def _job_ad_suffix(state: str) -> tuple[str, str] | None:
    if state == "Total":
        return "NATIONAL", "National"
    suffix = STATE_SUFFIX.get(state)
    if suffix is None:
        return None
    return suffix, state


def _salary_suffix(state: str, classification: str) -> tuple[str, str] | None:
    if state == "Total" and classification == "Total":
        return "NATIONAL", "National"
    if classification == "Total" and state != "Total":
        suffix = STATE_SUFFIX.get(state)
        if suffix is None:
            return None
        return suffix, state
    if state == "Total" and classification != "Total":
        return _industry_suffix(classification), classification
    # Unexpected cross-combo (state and classification both non-Total) --
    # not present in the published workbook (confirmed 2026-07-14); skip
    # defensively rather than raise, in case a future release adds one.
    return None


def _to_date(value) -> datetime.date:
    return value.date() if hasattr(value, "date") else value


def job_ad_rows_to_records(
    raw_rows: list[dict],
    *,
    since: datetime.date | None = None,
    until: datetime.date | None = None,
    now: datetime.datetime | None = None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    now = now or datetime.datetime.now(UTC)
    indicators: dict[str, IndicatorRow] = {}
    observations: list[ObservationRow] = []
    for row in raw_rows:
        mapped = _job_ad_suffix(row["state"])
        if mapped is None:
            continue
        suffix, label = mapped
        d = _to_date(row["date"])
        if since and d < since:
            continue
        if until and d > until:
            continue
        for variant, field, variant_label, is_sa in (
            ("INDEX", "ads_sa_index", "Seasonally Adjusted", True),
            ("INDEX_TREND", "ads_trend_index", "Trend", False),
        ):
            v = row.get(field)
            if v is None:
                continue
            imdr_code = f"SEEK.JOBADS.{variant}.{suffix}.AU"
            if imdr_code not in indicators:
                indicators[imdr_code] = IndicatorRow(
                    imdr_code=imdr_code,
                    vendor_name=VENDOR_NAME,
                    source_code=f"SEEK.JOBADS.{variant}.{suffix}",
                    display_name=(
                        f"SEEK Advertised Job Index — {label} "
                        f"({variant_label}, 2016=100)"
                    ),
                    unit="index",
                    frequency="MONTHLY",
                    country_iso="AU",
                    category="labour",
                    is_seasonally_adjusted=is_sa,
                )
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=d, vintage=0,
                release_date=now, value=float(v), ingested_at=now,
            ))
    return list(indicators.values()), observations


def salary_rows_to_records(
    raw_rows: list[dict],
    *,
    since: datetime.date | None = None,
    until: datetime.date | None = None,
    now: datetime.datetime | None = None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    now = now or datetime.datetime.now(UTC)
    indicators: dict[str, IndicatorRow] = {}
    observations: list[ObservationRow] = []
    for row in raw_rows:
        mapped = _salary_suffix(row["state"], row["classification"])
        if mapped is None:
            continue
        suffix, label = mapped
        d = _to_date(row["date"])
        if since and d < since:
            continue
        if until and d > until:
            continue
        for variant, field, variant_label, is_sa in (
            ("INDEX", "salary_sa_index", "Seasonally Adjusted", True),
            ("INDEX_TREND", "salary_trend_index", "Trend", False),
        ):
            v = row.get(field)
            if v is None:
                continue
            imdr_code = f"SEEK.SALARY.{variant}.{suffix}.AU"
            if imdr_code not in indicators:
                indicators[imdr_code] = IndicatorRow(
                    imdr_code=imdr_code,
                    vendor_name=VENDOR_NAME,
                    source_code=f"SEEK.SALARY.{variant}.{suffix}",
                    display_name=(
                        f"SEEK Advertised Salary Index — {label} "
                        f"({variant_label}, index)"
                    ),
                    unit="index",
                    frequency="MONTHLY",
                    country_iso="AU",
                    category="labour",
                    is_seasonally_adjusted=is_sa,
                )
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=d, vintage=0,
                release_date=now, value=float(v), ingested_at=now,
            ))
    return list(indicators.values()), observations


def build_rows(
    since: str | None = None,
    until: str | None = None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    lo = datetime.date.fromisoformat(since) if since else None
    hi = datetime.date.fromisoformat(until) if until else None

    with _make_client() as client:
        html = fetch_report_page_html(client)
        employment_url, salary_url = discover_download_urls(html)
        print(f"  employment download: {employment_url}")
        print(f"  salary download:     {salary_url}")

        employment_wb = _load_workbook(fetch_workbook_bytes(client, employment_url))
        salary_wb = _load_workbook(fetch_workbook_bytes(client, salary_url))

    job_ad_raw = parse_job_ad_sheet(employment_wb)
    salary_raw = parse_salary_sheet(salary_wb)

    job_ad_ind, job_ad_obs = job_ad_rows_to_records(job_ad_raw, since=lo, until=hi)
    salary_ind, salary_obs = salary_rows_to_records(salary_raw, since=lo, until=hi)

    print(f"  JOBADS  n_raw_rows={len(job_ad_raw)} n_indicators={len(job_ad_ind)} n_obs={len(job_ad_obs)}")
    print(f"  SALARY  n_raw_rows={len(salary_raw)} n_indicators={len(salary_ind)} n_obs={len(salary_obs)}")

    return job_ad_ind + salary_ind, job_ad_obs + salary_obs
