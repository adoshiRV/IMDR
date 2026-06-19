"""DPIIT 8-Core Industries Index — monthly level + YoY by sector.

Source: ``eaindustry.nic.in/eight_core_infra/Core_Industries_2011_12_*.xlsx``
(same site as WPI). Base 2011-12=100. Each release ships the full
back-history (Apr 2011 → latest) in two sheets — `Index` (levels) and
`Growth (%)` (YoY).

ICI is the sectoral output proxy that LEADS the IIP by ~10 days each
month — published last working day of the month for M-2 data, IIP
follows ~3 weeks later. 9 series × 2 metrics × ~180 mo ≈ 3,150 obs.

Cell mapping (see docs/admin/econ/india/in_coverage_plan.md):
  1.4 Macro core      — Overall index leads IIP
  Cluster 3           — sectoral activity
  2.1 Input costs     — Coal/Crude/NG/Petroleum upstream of WPI
"""
from __future__ import annotations

import datetime
import io
import re
import warnings

import httpx
from openpyxl import load_workbook

from imdr.domains.econ.schema import IndicatorRow, ObservationRow
from scripts.econ._runner import run_main

warnings.filterwarnings("ignore")

UTC = datetime.timezone.utc

_BASE = "https://eaindustry.nic.in"
_UA = {"User-Agent": "Mozilla/5.0", "Referer": _BASE + "/"}
_FILE_RE = re.compile(r"eight_core_infra/Core_Industries_2011_12_(\d+)\.xlsx")

_INDEX_STEMS: dict[str, tuple[str, str]] = {
    "Overall Index":         ("OVERALL",       "India 8-Core Industries — Overall Index"),
    "Index of Coal":         ("COAL",          "India 8-Core — Coal"),
    "Index of Crude Oil":    ("CRUDE_OIL",     "India 8-Core — Crude Oil"),
    "Index of Natural Gas":  ("NATURAL_GAS",   "India 8-Core — Natural Gas"),
    "Index of Petroleum":    ("PETROLEUM_REF", "India 8-Core — Petroleum Refinery"),
    "Index of Fertilize":    ("FERTILIZERS",   "India 8-Core — Fertilizers"),
    "Index of Steel":        ("STEEL",         "India 8-Core — Steel"),
    "Index of Cement":       ("CEMENT",        "India 8-Core — Cement"),
    "Index of Electricity":  ("ELECTRICITY",   "India 8-Core — Electricity"),
}
_GROWTH_STEMS: dict[str, tuple[str, str]] = {
    "Overall Growth":        ("OVERALL",       "India 8-Core Industries — Overall YoY %"),
    "Growth of Coal":        ("COAL",          "India 8-Core Coal YoY %"),
    "Growth of  Crude":      ("CRUDE_OIL",     "India 8-Core Crude Oil YoY %"),
    "Growth of  Natural":    ("NATURAL_GAS",   "India 8-Core Natural Gas YoY %"),
    "Growth of  Petroleum":  ("PETROLEUM_REF", "India 8-Core Petroleum Refinery YoY %"),
    "Growth of  Fertili":    ("FERTILIZERS",   "India 8-Core Fertilizers YoY %"),
    "Growth of  Steel":      ("STEEL",         "India 8-Core Steel YoY %"),
    "Growth of  Cement":     ("CEMENT",        "India 8-Core Cement YoY %"),
    "Growth of  Electric":   ("ELECTRICITY",   "India 8-Core Electricity YoY %"),
}


def _discover_url(client: httpx.Client) -> str:
    r = client.get(_BASE + "/", headers=_UA, timeout=30)
    r.raise_for_status()
    files = sorted(set(_FILE_RE.findall(r.text)))
    if not files:
        raise RuntimeError("no Core Industries XLSX link on DPIIT page")
    return f"{_BASE}/eight_core_infra/Core_Industries_2011_12_{files[-1]}.xlsx"


def _match_stem(label: str, m: dict[str, tuple[str, str]]) -> tuple[str, str] | None:
    if not label:
        return None
    for prefix, val in m.items():
        if str(label).startswith(prefix):
            return val
    return None


def _parse_sheet(
    ws, stems: dict[str, tuple[str, str]], metric: str, unit: str,
    now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen: set[str] = set()

    header = [str(c.value) if c.value is not None else "" for c in ws[1]]
    series_cols: list[tuple[int, str, str]] = []
    for j, h in enumerate(header):
        if j == 0:
            continue
        s = _match_stem(h.strip(), stems)
        if s is not None:
            stem, display = s
            series_cols.append((j, stem, display))
    if not series_cols:
        return indicators, observations

    for row in ws.iter_rows(min_row=2, values_only=True):
        cell0 = row[0]
        if not isinstance(cell0, (datetime.date, datetime.datetime)):
            continue
        obs_date = cell0.date() if isinstance(cell0, datetime.datetime) else cell0
        for col_j, stem, display in series_cols:
            v = row[col_j] if col_j < len(row) else None
            try:
                value = float(v)
            except (ValueError, TypeError):
                continue
            imdr_code = f"INDIA.CORE_IND.{stem}.{metric}.IN"
            if imdr_code not in seen:
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="DPIIT",
                    source_code=f"DPIIT/OEA/CORE_IND_2011_12/{stem}/{metric}",
                    display_name=display + (
                        " — index (2011-12=100)" if metric == "LEVEL" else " — YoY %"
                    ),
                    unit=unit, frequency="MONTHLY",
                    country_iso="IN", category="gdp",
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
                seen.add(imdr_code)
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=obs_date, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def run_fetch(
    since: str | None,
    until: str | None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    since_dt = datetime.date.fromisoformat(since) if since else None
    until_dt = datetime.date.fromisoformat(until) if until else None
    now = datetime.datetime.now(UTC)

    with httpx.Client(timeout=60, follow_redirects=True) as c:
        url = _discover_url(c)
        print(f"  latest 8-Core XLSX: {url}")
        rf = c.get(url, headers=_UA)
        rf.raise_for_status()
        blob = rf.content
    print(f"  {len(blob)} bytes")

    wb = load_workbook(io.BytesIO(blob), data_only=True)

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen: set[str] = set()

    for sheet, stems, metric, unit in (
        ("Index",      _INDEX_STEMS,  "LEVEL", "index"),
        ("Growth (%)", _GROWTH_STEMS, "YOY",   "pct"),
    ):
        if sheet not in wb.sheetnames:
            print(f"  skip — sheet missing: {sheet}")
            continue
        inds, obs = _parse_sheet(wb[sheet], stems, metric, unit, now)
        for i in inds:
            if i.imdr_code not in seen:
                indicators.append(i)
                seen.add(i.imdr_code)
        for o in obs:
            if since_dt and o.obs_date < since_dt:
                continue
            if until_dt and o.obs_date > until_dt:
                continue
            observations.append(o)
    return indicators, observations


def main() -> int:
    return run_main(vendor="dpiit", topic="core_industries",
                    fetch_fn=run_fetch,
                    description=__doc__.splitlines()[0] if __doc__ else "",
                    country_code="IN")


if __name__ == "__main__":
    import sys
    sys.exit(main())
