"""MOSPI Consumer Price Index — All-India + division CPI (2024 base).

Source: MOSPI "CPI for the month of ..." press release. Each release
ships an Annexure-I XLSX with Rural / Urban / Combined × Index + YoY
for the All-India headline and 12 sub-divisions (food/cloth/housing/...).

Each release adds one month. The 2024-base series begins April 2024;
historical 2012-base data lives in earlier releases and is NOT
splice-compatible (geographic axis was renumbered). Defer 2012-base
backfill to a separate stream.

This fetcher walks the listing API for the most-recent N CPI releases
(`MOSPI_CPI_MAX_RELEASES`, default 24 = 2 years) and emits 78 indicators
(13 divisions incl. headline × 3 geographies × 2 metrics) × N months
of observations.

Cell mapping (see docs/admin/econ/india/in_coverage_plan.md):
  1.2 Headline price       — INDIA.CPI.HEADLINE.{R,U,C}.{LEVEL,YOY}.IN
  1.3 Core / sub-divisions — division-level series
"""
from __future__ import annotations

import datetime
import io
import os
import re

import httpx
from openpyxl import load_workbook

from imdr.domains.econ.mospi import (
    HEADERS, list_releases, fetch_attachment,
)
from imdr.domains.econ.schema import IndicatorRow, ObservationRow
from scripts.econ._runner import run_main

UTC = datetime.timezone.utc

_DIVISIONS: dict[str, tuple[str, str]] = {
    "1":  ("FOOD_BEV",       "Food and beverages"),
    "2":  ("PAAN_TOB_INTOX", "Paan, tobacco and intoxicants"),
    "3":  ("CLOTH_FOOT",     "Clothing and footwear"),
    "4":  ("HOUSING",        "Housing, water, electricity, gas, etc."),
    "5":  ("FURNISH_HH",     "Furnishings, household equipment"),
    "6":  ("HEALTH",         "Health"),
    "7":  ("TRANSPORT",      "Transport"),
    "8":  ("INFO_COMMS",     "Information and communication"),
    "9":  ("RECREATION",     "Recreation, sport and culture"),
    "10": ("EDUCATION",      "Education services"),
    "11": ("RESTAURANTS",    "Restaurants and accommodation"),
    "13": ("PERSONAL_CARE",  "Personal care, social protection"),
}

_MONTH_RE = re.compile(r"month of (\w+),?\s*(\d{4})", re.I)


def _parse_release_month(title: str) -> datetime.date | None:
    m = _MONTH_RE.search(title or "")
    if not m:
        return None
    try:
        return datetime.datetime.strptime(
            f"{m.group(1)} {m.group(2)}", "%B %Y"
        ).date()
    except ValueError:
        return None


def _num(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "*", "-"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_annexure(
    blob: bytes, obs_date: datetime.date, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    wb = load_workbook(io.BytesIO(blob), data_only=True)
    if "Annexure-I" not in wb.sheetnames:
        return indicators, observations

    for row in wb["Annexure-I"].iter_rows(values_only=True):
        if not row or len(row) < 8:
            continue
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip().replace("\xa0", " ") if row[1] is not None else ""
        is_division = col0 in _DIVISIONS
        is_headline = "All India" in col1
        if not (is_division or is_headline):
            continue

        r_idx, u_idx, c_idx, r_yoy, u_yoy, c_yoy = (_num(row[i]) for i in range(2, 8))
        if is_headline:
            div_stem, div_name = "HEADLINE", "All India headline"
        else:
            div_stem, div_name = _DIVISIONS[col0]

        for geo, idx, yoy in (("R", r_idx, r_yoy),
                              ("U", u_idx, u_yoy),
                              ("C", c_idx, c_yoy)):
            geo_full = {"R": "Rural", "U": "Urban", "C": "Combined"}[geo]
            for metric, value, unit in (("LEVEL", idx, "index"),
                                         ("YOY",   yoy, "pct")):
                if value is None:
                    continue
                imdr_code = f"INDIA.CPI.{div_stem}.{geo}.{metric}.IN"
                display = (
                    f"India CPI {div_name} — {geo_full}, "
                    f"{'index (base 2024=100)' if metric == 'LEVEL' else 'YoY %'}"
                )
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="MOSPI",
                    source_code=f"MOSPI/Annexure-I/{div_stem}/{geo}/{metric}",
                    display_name=display, unit=unit, frequency="MONTHLY",
                    country_iso="IN", category="cpi",
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
                observations.append(ObservationRow(
                    imdr_code=imdr_code, obs_date=obs_date, vintage=0,
                    release_date=now, value=value, ingested_at=now,
                ))
    return indicators, observations


def run_fetch(
    since: str | None,
    until: str | None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    since_dt = datetime.date.fromisoformat(since) if since else None
    until_dt = datetime.date.fromisoformat(until) if until else None
    now = datetime.datetime.now(UTC)
    max_releases = int(os.environ.get("MOSPI_CPI_MAX_RELEASES", "24"))

    indicators_by_code: dict[str, IndicatorRow] = {}
    observations: list[ObservationRow] = []
    seen_obs: set[tuple[str, datetime.date]] = set()

    with httpx.Client(timeout=60, follow_redirects=True) as c:
        # listing API caps page_size at 50 — keep current page-1 behaviour;
        # 24-release window comfortably fits.
        releases = list_releases(c, "CPI for", page_size=50)
        # only keep items with an XLSX attachment
        releases = [r for r in releases
                    if (r.get("file_two") or {}).get("path", "").lower()
                       .endswith((".xlsx", ".xls"))]
        releases = releases[:max_releases]
        print(f"  {len(releases)} CPI releases to process")
        for rel in releases:
            obs_date = _parse_release_month(rel.get("title", ""))
            if obs_date is None:
                continue
            if since_dt and obs_date < since_dt:
                continue
            if until_dt and obs_date > until_dt:
                continue
            f2 = rel["file_two"]
            blob = fetch_attachment(c, f2["path"])
            inds, obs = _parse_annexure(blob, obs_date, now)
            for i in inds:
                indicators_by_code.setdefault(i.imdr_code, i)
            for o in obs:
                key = (o.imdr_code, o.obs_date)
                if key in seen_obs:
                    continue
                seen_obs.add(key)
                observations.append(o)
            print(f"    {obs_date}: +{len(obs)} obs")
    return list(indicators_by_code.values()), observations


def main() -> int:
    return run_main(vendor="mospi", topic="cpi",
                    fetch_fn=run_fetch,
                    description=__doc__.splitlines()[0] if __doc__ else "",
                    country_code="IN")


if __name__ == "__main__":
    import sys
    sys.exit(main())
