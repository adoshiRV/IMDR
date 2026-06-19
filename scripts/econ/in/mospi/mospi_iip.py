"""MOSPI Index of Industrial Production — monthly NIC + UBC headlines.

Source: MOSPI Quick Estimates of IIP press release (Base 2011-12 = 100).
Each release ships the full back-history (Apr 2012 → latest) as one
XLSX, so the latest item is sufficient for full coverage.

Emits 20 indicators × ~168 months ≈ 3,360 observations:

  4 sectoral (Mining, Manufacturing, Electricity, General) × LEVEL + YoY
  6 use-based categories                                  × LEVEL + YoY

NIC 2-digit industry-level series (~26 × 168 mo) deferred — not headline
macro.

Cell mapping (see docs/admin/econ/india/in_coverage_plan.md):
  2.1 Activity      — General / sectoral / UBC headlines
  2.2 Capex / inv   — UBC Capital + Infrastructure
"""
from __future__ import annotations

import datetime
import io
from openpyxl import load_workbook

import httpx

from imdr.domains.econ.mospi import latest_xlsx
from imdr.domains.econ.schema import IndicatorRow, ObservationRow
from scripts.econ._runner import run_main

UTC = datetime.timezone.utc

_SECTORAL: dict[str, tuple[str, str]] = {
    "Mining":        ("MINING",        "India IIP — Mining sector (NSO Base 2011-12)"),
    "Manufacturing": ("MANUFACTURING", "India IIP — Manufacturing sector (NSO Base 2011-12)"),
    "Electricity":   ("ELECTRICITY",   "India IIP — Electricity sector (NSO Base 2011-12)"),
    "General":       ("GENERAL",       "India IIP — General (Headline, NSO Base 2011-12)"),
}
_UBC: dict[str, tuple[str, str]] = {
    "primary goods":                       ("UBC_PRIMARY",      "India IIP — Primary goods (UBC, NSO Base 2011-12)"),
    "capital goods":                       ("UBC_CAPITAL",      "India IIP — Capital goods (UBC, NSO Base 2011-12)"),
    "intermediate goods":                  ("UBC_INTERMEDIATE", "India IIP — Intermediate goods (UBC, NSO Base 2011-12)"),
    "infrastructure/ construction goods":  ("UBC_INFRA_CONST",  "India IIP — Infrastructure / construction goods (UBC, NSO Base 2011-12)"),
    "consumer durables":                   ("UBC_CONS_DUR",     "India IIP — Consumer durables (UBC, NSO Base 2011-12)"),
    "consumer non-durables":               ("UBC_CONS_NDUR",    "India IIP — Consumer non-durables (UBC, NSO Base 2011-12)"),
}


def _parse_sheet(
    wb, sheet_name: str, label_map: dict[str, tuple[str, str]],
    end_month: datetime.date, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for r in rows:
        if r and (r[0] in ("NIC 2008", "Use-based category", "UBC")):
            header_row = r
            break
    if header_row is None:
        return indicators, observations

    data_start = None
    for j, val in enumerate(header_row):
        if val is None:
            continue
        if str(val).lower() in ("nic 2008", "description", "weights", "weight",
                                 "use-based category", "ubc"):
            continue
        data_start = j
        break
    if data_start is None:
        return indicators, observations

    n_cols = 0
    for j in range(data_start, len(header_row)):
        if header_row[j] is None:
            break
        n_cols += 1

    months: list[datetime.date] = []
    y, m = end_month.year, end_month.month
    for _ in range(n_cols):
        months.append(datetime.date(y, m, 1))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    months.reverse()
    print(f"  {sheet_name}: data_start={data_start}, {n_cols} months "
          f"({months[0]} → {months[-1]})")

    in_growth = False
    for r in rows:
        if not r:
            continue
        col_a = r[0]
        if col_a and "growth" in str(col_a).lower():
            in_growth = True
            continue

        if sheet_name.startswith("NIC"):
            key = str(col_a).strip() if col_a is not None else ""
        else:
            key = str(col_a).strip().lower() if col_a is not None else ""
        if key not in label_map:
            continue
        stem, display = label_map[key]

        metric = "YOY" if in_growth else "LEVEL"
        unit = "pct" if in_growth else "index"
        imdr_code = f"INDIA.IIP.{stem}.{metric}.IN"
        display_full = (
            f"{display} — {'YoY % growth' if in_growth else 'index level'}"
        )
        if imdr_code not in seen_codes:
            indicators.append(IndicatorRow(
                imdr_code=imdr_code, vendor_name="MOSPI",
                source_code=f"MOSPI/IIP/{sheet_name}/{stem}/{metric}",
                display_name=display_full,
                unit=unit, frequency="MONTHLY",
                country_iso="IN", category="gdp",
                is_seasonally_adjusted=False, bbg_ticker=None,
            ))
            seen_codes.add(imdr_code)
        for k, month_date in enumerate(months):
            v = r[data_start + k] if data_start + k < len(r) else None
            try:
                value = float(v)
            except (ValueError, TypeError):
                continue
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=month_date, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def run_fetch(
    since: str | None,
    until: str | None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    since_dt = datetime.date.fromisoformat(since) if since else None
    until_dt = datetime.date.fromisoformat(until) if until else None
    now = datetime.datetime.now(UTC)

    with httpx.Client(timeout=60, follow_redirects=True) as c:
        print("  fetching latest IIP release...")
        title, end_month, blob = latest_xlsx(c, "Quick Estimates of IIP")
        print(f"    {title}")
        print(f"    end_month: {end_month}")

    wb = load_workbook(io.BytesIO(blob), data_only=True)

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for sheet, label_map in (("NIC 2d, sectoral monthly", _SECTORAL),
                              ("UBC monthly", _UBC)):
        if sheet not in wb.sheetnames:
            print(f"  skip — sheet missing: {sheet}")
            continue
        inds, obs = _parse_sheet(wb, sheet, label_map, end_month, now)
        for i in inds:
            if i.imdr_code not in seen_codes:
                indicators.append(i)
                seen_codes.add(i.imdr_code)
        for o in obs:
            if since_dt and o.obs_date < since_dt:
                continue
            if until_dt and o.obs_date > until_dt:
                continue
            observations.append(o)
    return indicators, observations


def main() -> int:
    return run_main(vendor="mospi", topic="iip",
                    fetch_fn=run_fetch,
                    description=__doc__.splitlines()[0] if __doc__ else "",
                    country_code="IN")


if __name__ == "__main__":
    import sys
    sys.exit(main())
