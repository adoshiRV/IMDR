"""RBI Monthly Bulletin — XLSX table fetcher (TSPD-cleared headed Chrome).

The RBI Bulletin publishes ~200 statistical tables once a month at
`rbidocs.rbi.org.in/rdocs/Bulletin/DOCs/*.XLSX`. The bulletin XLSX files
are HTML-tables-with-XLSX-extension behind **Akamai TSPD**
bot-protection. Plain httpx and headless Playwright are blocked
(verified 2026-06-11). HEADED Chrome with active JS execution clears
TSPD's challenge naturally; the file downloads as a normal browser
download.

Each XLSX is a **single-month snapshot** — typically 4-6 monthly
periods (or ~28 days for daily series). Production strategy: monthly
tick that appends new periods to existing indicators (MERGE on PK is
idempotent). For deep history → RBI Handbook of Statistics (separate
annual XLSX dump, not yet probed).

Reuses the headed-Chrome download + parser stack from
[`playground/econ/rbi/fetch_bulletin.py`](../../../rbi/fetch_bulletin.py).
Ported here under `scripts/econ/in/rbi/` per country-first layout +
adapted to the prod `imdr.domains.econ.schema` + `run_main` shape.

Run (headed window pops for ~30s per table):
    python -m scripts.econ.in.rbi.rbi_bulletin --no-load

Always-headed. If you need to verify headless-is-blocked behaviour,
use the standalone scaffold at `playground/econ/rbi/fetch_bulletin.py`
which carries its own argparse.
"""
from __future__ import annotations

import datetime
import io
import re
import sys
import time
from pathlib import Path

import openpyxl

from imdr.domains.econ.schema import IndicatorRow, ObservationRow
from scripts.econ._runner import run_main

UTC = datetime.timezone.utc

# Runtime dirs anchored under the repo data tree (country-first).
# parents[0]=rbi, [1]=in, [2]=econ, [3]=scripts, [4]=repo root
_REPO_ROOT = Path(__file__).resolve().parents[4]
PROFILE = _REPO_ROOT / "data" / "econ" / "in" / "rbi" / "_profile"
DL_DIR  = _REPO_ROOT / "data" / "econ" / "in" / "rbi" / "_downloads"

BULLETIN_LANDING = "https://www.rbi.org.in/Scripts/BS_ViewBulletin.aspx"

# Filename prefix is "<num>T_BULL..." e.g. 34T_BULL..., 19CT_BULL...
_TABLE_NUM_RE = re.compile(r"/(\d+[A-Z]?)T_BULL", re.IGNORECASE)
_XLSX_HREF_RE = re.compile(r"rdocs/Bulletin/DOCs/.+\.XLSX", re.IGNORECASE)

# ---------------------------------------------------------------------------
# PRIORITY_TARGETS — keyed by table number, NOT by hard-coded URL.
# Discovery step (run once per session in a headed Chrome window) maps
# each number → current month's URL via _discover_urls().
# ---------------------------------------------------------------------------
PRIORITY_TARGETS: list[dict] = [
    # --- existing tables (unchanged semantics) ---
    {
        "name": "cpi_combined",
        "table_num": "19C",
        "imdr_prefix": "INDIA.RBI_BULLETIN.CPI_COMBINED",
        "category": "cpi", "frequency": "MONTHLY",
        "description": "RBI Bulletin T19(C) — CPI Combined by Division (Base 2024=100)",
        "parser": "parse_cpi_combined_19c",
    },
    {
        "name": "call_money_rates",
        "table_num": "27",
        "imdr_prefix": "INDIA.RBI_BULLETIN.CALL_MONEY",
        "category": "rates", "frequency": "DAILY",
        "description": "RBI Bulletin T27 — Daily Call Money Rates",
        "parser": "parse_call_money_27",
    },
    {
        "name": "iip_23",
        "table_num": "23",
        "imdr_prefix": "INDIA.RBI_BULLETIN.IIP",
        "category": "gdp", "frequency": "MONTHLY",
        "description": "RBI Bulletin T23 — Index of Industrial Production (Base 2011-12=100)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Industry", "data_first_col": 3, "unit": "index"},
    },
    {
        "name": "money_stock_6",
        "table_num": "6",
        "imdr_prefix": "INDIA.RBI_BULLETIN.MONEY_STOCK",
        "category": "liquidity", "frequency": "WEEKLY",
        "description": "RBI Bulletin T6 — Money Stock Measures M0/M1/M3 (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "inr_cr"},
    },
    {
        "name": "reserve_money_11",
        "table_num": "11",
        "imdr_prefix": "INDIA.RBI_BULLETIN.RESERVE_MONEY",
        "category": "cb_balance_sheet", "frequency": "WEEKLY",
        "description": "RBI Bulletin T11 — Reserve Money components & sources (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "inr_cr"},
    },
    {
        "name": "wpi_22",
        "table_num": "22",
        "imdr_prefix": "INDIA.RBI_BULLETIN.WPI",
        "category": "cpi", "frequency": "MONTHLY",
        "description": "RBI Bulletin T22 — Wholesale Price Index (Base 2011-12=100)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Commodities", "data_first_col": 3, "unit": "index"},
    },
    {
        "name": "rbi_bs_2",
        "table_num": "2",
        "imdr_prefix": "INDIA.RBI_BULLETIN.RBI_BS",
        "category": "cb_balance_sheet", "frequency": "WEEKLY",
        "description": "RBI Bulletin T2 — RBI Liabilities and Assets (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "inr_cr"},
    },
    {
        "name": "fx_reserves_33",
        "table_num": "33",
        "imdr_prefix": "INDIA.RBI_BULLETIN.FX_RESERVES",
        "category": "cb_balance_sheet", "frequency": "WEEKLY",
        "description": "RBI Bulletin T33 — Foreign Exchange Reserves (dual-unit)",
        "parser": "parse_dual_unit",
        "dual": {"data_first_col": 3},
    },
    {
        "name": "foreign_trade_32",
        "table_num": "32",
        "imdr_prefix": "INDIA.RBI_BULLETIN.FOREIGN_TRADE",
        "category": "bop", "frequency": "MONTHLY",
        "description": "RBI Bulletin T32 — Foreign Trade (dual-unit Exports/Imports/Oil/Non-oil)",
        "parser": "parse_dual_unit",
        "dual": {"data_first_col": 3},
    },
    {
        "name": "bop_40",
        "table_num": "40",
        "imdr_prefix": "INDIA.RBI_BULLETIN.BOP",
        "category": "bop", "frequency": "QUARTERLY",
        "description": "RBI Bulletin T40 — India's Overall Balance of Payments (USD Million)",
        "parser": "parse_bop",
    },
    # --- NEW tables (added 2026-06-18) ---
    {
        "name": "nri_deposits_34",
        "table_num": "34",
        "imdr_prefix": "INDIA.RBI_BULLETIN.NRI_DEPOSITS",
        "category": "bop", "frequency": "MONTHLY",
        "description": "RBI Bulletin T34 — Non-Resident Deposits FCNR(B)/NR(E)RA/NRO (USD Million)",
        "parser": "parse_nri_deposits_34",
    },
    {
        "name": "foreign_investment_35",
        "table_num": "35",
        "imdr_prefix": "INDIA.RBI_BULLETIN.FOREIGN_INVESTMENT",
        "category": "bop", "frequency": "MONTHLY",
        "description": "RBI Bulletin T35 — Foreign Investment Inflows (USD Million)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "usd_mn"},
    },
    {
        "name": "lrs_remittances_36",
        "table_num": "36",
        "imdr_prefix": "INDIA.RBI_BULLETIN.LRS_REMITTANCES",
        "category": "bop", "frequency": "MONTHLY",
        "description": "RBI Bulletin T36 — Outward Remittances under LRS (USD Million)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "usd_mn"},
    },
    {
        "name": "fx_turnover_30",
        "table_num": "30",
        "imdr_prefix": "INDIA.RBI_BULLETIN.FX_TURNOVER",
        "category": "fx", "frequency": "WEEKLY",
        "description": "RBI Bulletin T30 — Average Daily Turnover in Select Financial Markets (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {"header_match": "Item", "data_first_col": 2, "unit": "inr_cr"},
    },
    {
        "name": "tbill_ownership_25",
        "table_num": "25",
        "imdr_prefix": "INDIA.RBI_BULLETIN.TBILL_OWNERSHIP",
        "category": "rates", "frequency": "WEEKLY",
        "description": "RBI Bulletin T25 — Treasury Bills Ownership Pattern (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {
            "header_match": "Item", "data_first_col": 2, "unit": "inr_cr",
            "carry_section": True,
        },
    },
    {
        "name": "ecb_38",
        "table_num": "38",
        "imdr_prefix": "INDIA.RBI_BULLETIN.ECB",
        "category": "bop", "frequency": "MONTHLY",
        "description": "RBI Bulletin T38 — External Commercial Borrowings Registrations (USD Million)",
        "parser": "parse_wide_table",
        "wide": {
            "header_match": "Item", "data_first_col": 2, "unit": "usd_mn",
            "carry_section": True,
        },
    },
    {
        "name": "rbi_standing_5",
        "table_num": "5",
        "imdr_prefix": "INDIA.RBI_BULLETIN.RBI_STANDING",
        "category": "cb_balance_sheet", "frequency": "WEEKLY",
        "description": "RBI Bulletin T5 — RBI Standing Facilities (INR Crore)",
        "parser": "parse_wide_table",
        "wide": {
            "header_match": "Item", "data_first_col": 2, "unit": "inr_cr",
            "carry_section": True,
        },
    },
    {
        "name": "cd_28",
        "table_num": "28",
        "imdr_prefix": "INDIA.RBI_BULLETIN.CD",
        "category": "rates", "frequency": "WEEKLY",
        "description": "RBI Bulletin T28 — Certificates of Deposit (INR Crore / pct)",
        "parser": "parse_cd_cp",
    },
    {
        "name": "cp_29",
        "table_num": "29",
        "imdr_prefix": "INDIA.RBI_BULLETIN.CP",
        "category": "rates", "frequency": "WEEKLY",
        "description": "RBI Bulletin T29 — Commercial Paper (INR Crore / pct)",
        "parser": "parse_cd_cp",
    },
    {
        "name": "laf_3",
        "table_num": "3",
        "imdr_prefix": "INDIA.RBI_BULLETIN.LAF",
        "category": "cb_balance_sheet", "frequency": "DAILY",
        "description": "RBI Bulletin T3 — Liquidity Operations by RBI (INR Crore)",
        "parser": "parse_date_rows",
    },
    {
        "name": "iip_44",
        "table_num": "44",
        "imdr_prefix": "INDIA.RBI_BULLETIN.IIP_INTL",
        "category": "bop", "frequency": "QUARTERLY",
        "description": "RBI Bulletin T44 — International Investment Position (USD Million)",
        "parser": "parse_iip_assets_liab",
    },
    {
        "name": "tbill_auctions_26",
        "table_num": "26",
        "imdr_prefix": "INDIA.RBI_BULLETIN.TBILL_AUCTIONS",
        "category": "rates", "frequency": "WEEKLY",
        "description": "RBI Bulletin T26 — Auctions of Treasury Bills (INR Crore)",
        "parser": "parse_tbill_auctions_26",
    },
]


# ---------------------------------------------------------------------------
# URL auto-discovery (replaces hard-coded per-month hash URLs)
# ---------------------------------------------------------------------------

def _scrape_xlsx_links(page) -> dict[str, str]:
    """Scrape the bulletin index page and return {table_num: url} for all XLSX links."""
    anchors = page.eval_on_selector_all(
        "a",
        "els => els.map(a => ({href: a.href, text: (a.textContent||'').trim()}))",
    )
    result: dict[str, str] = {}
    seen: set[str] = set()
    for a in anchors:
        href = a.get("href") or ""
        if not _XLSX_HREF_RE.search(href):
            continue
        if href in seen:
            continue
        seen.add(href)
        m = _TABLE_NUM_RE.search(href)
        if m:
            num = m.group(1).upper()
            result[num] = href
    return result


def _discover_urls(page, wanted_nums: list[str]) -> dict[str, str]:
    """Return {table_num: url} for the given table numbers from the live index."""
    print(f"  discovering URLs from {BULLETIN_LANDING}")
    try:
        page.goto(BULLETIN_LANDING, timeout=45000, wait_until="domcontentloaded")
        page.wait_for_timeout(4000)
    except Exception as e:
        print(f"    discovery warning: {e}")

    url_map = _scrape_xlsx_links(page)
    if not url_map:
        print("    no XLSX links found on landing; trying Current-Statistics sub-page")
        for sel in (
            "a:has-text('Current Statistics')",
            "a:has-text('CURRENT STATISTICS')",
            "a:has-text('Statistics')",
        ):
            try:
                link = page.locator(sel).first
                if link.count():
                    link.click(timeout=8000)
                    page.wait_for_timeout(4000)
                    break
            except Exception:
                continue
        url_map = _scrape_xlsx_links(page)

    found = {n: url_map[n] for n in wanted_nums if n in url_map}
    missing = [n for n in wanted_nums if n not in url_map]
    print(f"    {len(found)}/{len(wanted_nums)} table URLs resolved; "
          f"{len(missing)} missing: {missing or 'none'}")
    return found


# ---------------------------------------------------------------------------
# Headed Chrome download (TSPD-cleared)
# ---------------------------------------------------------------------------

def _download_via_headed(targets: list[dict], *, headless: bool = False) -> list[Path]:
    from playwright.sync_api import sync_playwright

    PROFILE.mkdir(parents=True, exist_ok=True)
    DL_DIR.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []
    print(f"  launching {'HEADLESS (will fail)' if headless else 'headed'} Chrome")
    with sync_playwright() as pw:
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE),
            channel="chrome", headless=headless,
            accept_downloads=True,
            viewport={"width": 1280, "height": 900},
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()

        # --- discover current URLs for all wanted table numbers ---
        wanted_nums = [t["table_num"].upper() for t in targets]
        url_map = _discover_urls(page, wanted_nums)

        for i, t in enumerate(targets, 1):
            num = t["table_num"].upper()
            url = url_map.get(num)
            if not url:
                print(f"  [{i}/{len(targets)}] {t['name']} (T{num}) — URL not found in index, skip")
                continue
            outpath = DL_DIR / f"{t['name']}.xlsx"
            print(f"  [{i}/{len(targets)}] {t['name']} (T{num})")
            try:
                with page.expect_download(timeout=60000) as dl_info:
                    try:
                        page.goto(url, timeout=30000)
                    except Exception:
                        pass  # download triggers a Page.goto error we swallow
                download = dl_info.value
                download.save_as(str(outpath))
                head = outpath.read_bytes()[:4]
                if head.startswith(b"PK\x03\x04"):
                    size = outpath.stat().st_size
                    print(f"    OK  {size:,}B (XLSX)")
                    saved.append(outpath)
                elif head.startswith(b"<!DO") or head.startswith(b"<htm"):
                    print(f"    BLOCKED: TSPD challenge")
                    outpath.unlink(missing_ok=True)
                else:
                    print(f"    UNKNOWN format head={head!r}")
            except Exception as e:
                print(f"    FAIL: {type(e).__name__}: {str(e)[:100]}")
            time.sleep(2)
        ctx.close()
    return saved


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_sheet(path: Path, sheet_idx: int = 0) -> list[list[str]]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[sheet_idx]]
    rows: list[list[str]] = []
    for raw in ws.iter_rows(values_only=True):
        row: list[str] = []
        for v in raw:
            if v is None:
                row.append("")
            elif isinstance(v, datetime.datetime):
                row.append(v.date().isoformat())
            elif isinstance(v, datetime.date):
                row.append(v.isoformat())
            else:
                row.append(str(v).strip())
        while row and row[-1] == "":
            row.pop()
        if row:
            rows.append(row)
    wb.close()
    return rows


_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
    "%d %b %Y", "%d-%b-%Y", "%d-%b-%y",
    "%B %d, %Y", "%b %d, %Y",
    "%b. %d, %Y",           # RBI Bulletin T3: "Mar. 1, 2026"
    "%b %Y", "%b. %Y",      # "Mar 2026" and "Mar. 2026"
    "%B %Y", "%b-%y", "%B-%y", "%Y-%m", "%Y",
)


def _parse_date(s: str) -> datetime.date | None:
    s = (s or "").strip()
    if not s:
        return None
    s = re.sub(r"\s*\(P\)\s*$", "", s).strip().rstrip("*").strip()
    m = re.match(r"^\s*(\d{4})\s*-\s*\d{2,4}\s*$", s)
    if m:
        return datetime.date(int(m.group(1)), 4, 1)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _col_periods_from_header_block(
    rows: list[list[str]], header_row_idx: int,
    data_first_col: int, n_header_rows: int = 4,
    carry_year_across_cols: bool = False,
) -> dict[int, datetime.date]:
    """Shared period-column resolver used by parse_wide_table and variants.

    Walks up to `n_header_rows` rows starting at `header_row_idx`, column
    by column from `data_first_col`, and resolves the most specific date for
    each column. Returns {col_index: date}.

    carry_year_across_cols: when True, a year token seen in any earlier
    column of the same header row is carried forward to later columns
    that have only a month token. Needed for T34-style tables where the
    year "2026" appears in col N and the month "Mar. (P)" is in col N+1
    with an empty year cell.
    """
    max_col = max(
        (len(rows[header_row_idx + k])
         for k in range(n_header_rows)
         if header_row_idx + k < len(rows)),
        default=0,
    )
    col_period: dict[int, datetime.date] = {}

    if carry_year_across_cols:
        # Build a per-row, per-column array of cells.
        # Then pass left-to-right tracking the most recent year seen on each row.
        hdr: list[list[str]] = []
        for k in range(n_header_rows):
            r = rows[header_row_idx + k] if header_row_idx + k < len(rows) else []
            hdr.append(r)

        # row_year[k] will hold the running last_year scanned left-to-right on row k
        row_year: list[str | None] = [None] * n_header_rows

        for j in range(data_first_col, max_col):
            last_dt: datetime.date | None = None
            for k in range(n_header_rows):
                r = hdr[k]
                cell = (r[j] if j < len(r) else "").strip()
                if not cell:
                    continue
                if re.match(r"^\d{4}$", cell):
                    row_year[k] = cell
                    continue
                d = _parse_date(cell)
                if d is not None:
                    last_dt = d
                    continue
                yr = row_year[k]
                if yr:
                    t = re.sub(r"\s*\([PR]\)\s*$", "", cell).strip().rstrip(".").strip()
                    for fmt in ("%b %d %Y", "%b. %d %Y", "%b %Y", "%b. %Y",
                                 "%B %Y", "%B %d %Y"):
                        try:
                            last_dt = datetime.datetime.strptime(
                                f"{t} {yr}", fmt).date()
                            break
                        except ValueError:
                            continue
            if last_dt is not None:
                col_period[j] = last_dt
        return col_period

    for j in range(data_first_col, max_col):
        last_dt: datetime.date | None = None
        last_year: str | None = None
        for k in range(n_header_rows):
            r = rows[header_row_idx + k] if header_row_idx + k < len(rows) else []
            cell = (r[j] if j < len(r) else "").strip()
            if not cell:
                continue
            if re.match(r"^\d{4}$", cell):
                last_year = cell
                continue
            d = _parse_date(cell)
            if d is not None:
                last_dt = d
                continue
            if last_year:
                t = re.sub(r"\s*\([PR]\)\s*$", "", cell).strip().rstrip(".").strip()
                for fmt in ("%b %d %Y", "%b. %d %Y", "%b %Y", "%b. %Y",
                             "%B %Y", "%B %d %Y"):
                    try:
                        last_dt = datetime.datetime.strptime(
                            f"{t} {last_year}", fmt).date()
                        break
                    except ValueError:
                        continue
        if last_dt is not None:
            col_period[j] = last_dt
    return col_period


# ---------------------------------------------------------------------------
# Per-table parsers
# ---------------------------------------------------------------------------

def parse_call_money_27(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    base = target["imdr_prefix"]
    inds = [
        IndicatorRow(
            imdr_code=f"{base}.WAVG.IN", vendor_name="RBI",
            source_code="bulletin/T27/wavg",
            display_name="India Call Money — Weighted Average (RBI Bulletin T27)",
            unit="pct", frequency="DAILY", country_iso="IN",
            category="rates", is_seasonally_adjusted=False, bbg_ticker=None,
        ),
        IndicatorRow(
            imdr_code=f"{base}.RANGE_LO.IN", vendor_name="RBI",
            source_code="bulletin/T27/range_low",
            display_name="India Call Money — Range Low (RBI Bulletin T27)",
            unit="pct", frequency="DAILY", country_iso="IN",
            category="rates", is_seasonally_adjusted=False, bbg_ticker=None,
        ),
        IndicatorRow(
            imdr_code=f"{base}.RANGE_HI.IN", vendor_name="RBI",
            source_code="bulletin/T27/range_high",
            display_name="India Call Money — Range High (RBI Bulletin T27)",
            unit="pct", frequency="DAILY", country_iso="IN",
            category="rates", is_seasonally_adjusted=False, bbg_ticker=None,
        ),
    ]
    obs: list[ObservationRow] = []
    for r in rows:
        if len(r) < 4:
            continue
        d = _parse_date(r[1])
        if d is None:
            continue
        try:
            wavg = float((r[3] or "").strip().replace(",", ""))
        except ValueError:
            wavg = None
        lo = hi = None
        m = re.match(r"^\s*([\d.]+)\s*-\s*([\d.]+)\s*$", (r[2] or "").strip())
        if m:
            try:
                lo, hi = float(m.group(1)), float(m.group(2))
            except ValueError:
                pass
        for code, value in ((inds[0].imdr_code, wavg),
                             (inds[1].imdr_code, lo),
                             (inds[2].imdr_code, hi)):
            if value is None:
                continue
            obs.append(ObservationRow(
                imdr_code=code, obs_date=d, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return inds, obs


def parse_cpi_combined_19c(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    base = target["imdr_prefix"]

    # Detect layout rows. Bulletin T19C is wide: rows 2-4 are headers,
    # rows 5-N are divisions.
    measure_row_idx = period_row_idx = data_start_idx = None
    for i, r in enumerate(rows):
        if measure_row_idx is None:
            joined = " | ".join(str(c) for c in r)
            if "Index" in joined and "Inflation" in joined:
                measure_row_idx = i
                continue
        if measure_row_idx is not None and period_row_idx is None:
            if any(_parse_date(str(c)) for c in r[2:]):
                period_row_idx = i
                continue
        if period_row_idx is not None and data_start_idx is None:
            label = (r[1] or "").strip() if len(r) > 1 else ""
            if label and (re.match(r"^\d+\.", label) or label.startswith("All India")):
                data_start_idx = i
                break
    if measure_row_idx is None or period_row_idx is None or data_start_idx is None:
        return [], []

    measure_row = rows[measure_row_idx]
    period_row = rows[period_row_idx]
    col_measure: dict[int, str] = {}
    current = ""
    for ci, v in enumerate(measure_row):
        lbl = (v or "").strip()
        if lbl == "Index":
            current = "INDEX"
        elif lbl == "Inflation (y-o-y, per cent)":
            current = "YOY"
        col_measure[ci] = current
    col_period: dict[int, datetime.date | None] = {}
    for ci, raw in enumerate(period_row):
        if ci < 3 or raw is None or raw == "":
            continue
        s = str(raw).strip()
        d = _parse_date(s) or _parse_date(s.replace("(P)", "").strip())
        col_period[ci] = d

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for r in rows[data_start_idx:]:
        if not r or len(r) < 3:
            continue
        division = (r[1] or "").strip()
        if not division or division.lower().startswith(("memo", "p:", "sources")):
            continue
        slug = re.sub(r"^\d+\.\s*", "", division)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug).strip("_").upper()[:50] or "DIV"
        for measure in ("INDEX", "YOY"):
            imdr_code = f"{base}.{slug}.{measure}.IN"
            if imdr_code not in seen_codes:
                seen_codes.add(imdr_code)
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="RBI",
                    source_code=f"bulletin/T19C/{slug}/{measure}",
                    display_name=(
                        f"India CPI Combined {division} — {measure} "
                        f"(RBI Bulletin T19C, Base 2024=100)"
                    )[:255],
                    unit="index" if measure == "INDEX" else "pct",
                    frequency="MONTHLY", country_iso="IN", category="cpi",
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
        for ci, period in col_period.items():
            if period is None or ci >= len(r):
                continue
            raw = r[ci]
            cell = (raw if isinstance(raw, str) else str(raw)
                    ).replace(",", "").replace("–", "").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            measure = col_measure.get(ci, "")
            if not measure:
                continue
            imdr_code = f"{base}.{slug}.{measure}.IN"
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=period, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_wide_table(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """Generic parser for `item-x-period-columns` RBI Bulletin tables.

    Used for IIP T23, Money Stock T6, Reserve Money T11, NEER/REER T37,
    and the new additive tables T35/T36/T30/T25/T38/T5.

    Configured via ``target["wide"]``:

      header_match   substring in the header row (e.g. ``"Item"``,
                     ``"Industry"``) — used to locate the multi-row
                     period header.
      data_first_col 0-indexed column where period data starts (the
                     item label is always col 1; col 2 is weight if
                     present).
      unit           IMDR unit code for all emitted indicators.
      carry_section  If True, value-less rows (section headers like
                     "1. 91-day" or "2. Export Credit") carry their label
                     forward as a prefix for the following sub-rows. Used
                     by T25 (T-bill ownership by tenor), T38 (ECB by
                     permission type), T5 (standing facilities by type).
    """
    cfg = target.get("wide", {})
    header_match = cfg.get("header_match", "Item")
    data_first_col = int(cfg.get("data_first_col", 2))
    unit = cfg.get("unit", "index")
    carry_section = bool(cfg.get("carry_section", False))
    base = target["imdr_prefix"]

    # Locate the row whose col-1 cell contains the header match.
    header_row_idx = None
    for i, r in enumerate(rows):
        if len(r) >= 2 and header_match in (r[1] or ""):
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], []

    col_period = _col_periods_from_header_block(
        rows, header_row_idx, data_first_col, n_header_rows=4,
    )
    if not col_period:
        return [], []

    # Find first data row dynamically — walk forward from header_row_idx
    # until we hit a row that has a non-empty value in at least 2 of the
    # mapped period columns. This replaces the original hard-coded
    # `header_row_idx + 4` offset, which under-counted T37 NEER/REER
    # because that table's data row sits at +5 (extra FY-tag row).
    data_start = header_row_idx + 1
    period_cols = list(col_period.keys())
    for di in range(header_row_idx + 1, min(header_row_idx + 8, len(rows))):
        r = rows[di]
        hits = sum(
            1 for ci in period_cols
            if ci < len(r) and (r[ci] or "").strip() not in ("", "1", "2", "3", "4", "5", "6", "7", "8", "9")
            and re.match(r"^[-\d.,]+$", (r[ci] or "").strip())
        )
        if hits >= 2:
            data_start = di
            break

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()
    current_section: str = ""

    # When carry_section is True we must start from header_row_idx+1 so that
    # section header rows that appear before the first data row (e.g. "1. 91-day"
    # in T25) are seen and captured in current_section. Otherwise data_start is
    # sufficient and avoids re-processing header/number rows.
    scan_start = header_row_idx + 1 if carry_section else data_start

    for r in rows[scan_start:]:
        if not r or len(r) < data_first_col + 1:
            # carry_section: a short row with just a label is a section header
            if carry_section and r and len(r) >= 2:
                lbl = (r[1] or "").strip().replace("\xa0", " ")
                if lbl and not any(lbl.lower().startswith(p)
                                   for p in ("source", "note", "p :", "p:")):
                    current_section = lbl
            continue
        label = (r[1] or "").strip().replace("\xa0", " ")
        if not label:
            continue
        if any(label.lower().startswith(p) for p in ("source", "note", "p :", "p:")):
            break

        # Section header detection: label present but no numeric value in data cols.
        first_val = (r[data_first_col] if data_first_col < len(r) else "").strip()
        if not first_val:
            if carry_section:
                current_section = label
            continue

        full_label = (
            f"{current_section} — {label}" if carry_section and current_section
            else label
        )

        # Build slug from full_label (strip leading numbering, normalise)
        slug_raw = re.sub(r"^[\d.]+\s*", "", full_label)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug_raw).strip("_").upper()[:50] or "ROW"
        imdr_code = f"{base}.{slug}.IN"
        if imdr_code not in seen_codes:
            seen_codes.add(imdr_code)
            indicators.append(IndicatorRow(
                imdr_code=imdr_code, vendor_name="RBI",
                source_code=f"bulletin/{target['name']}/{slug}",
                display_name=f"{target['description']} — {full_label}"[:255],
                unit=unit, frequency=target["frequency"], country_iso="IN",
                category=target["category"],
                is_seasonally_adjusted=False, bbg_ticker=None,
            ))
        for ci, period in col_period.items():
            if ci >= len(r):
                continue
            cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
            cell = cell.replace(",", "").replace("–", "").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=period, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_dual_unit(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """Dual-unit-row parser for T32 Foreign Trade + T33 FX Reserves.

    Layout (T33 FX Reserves shown):
      R2  | Item        | Unit         | <FY-or-year>   | <year>       …
      R3  |             |              | <month-token>  | <month-token>…
      R5  | 1. Total    | ₹ Crore      | 5797792 …
      R6  |             | US $ Million | 686064 …

    Each "item × unit" pair becomes one indicator. Item label carries
    forward across sub-rows with blank col-1.
    """
    cfg = target.get("dual", {})
    data_first_col = int(cfg.get("data_first_col", 3))
    unit_to_imdr = {
        "₹ Crore": "inr_cr",
        "INR Crore": "inr_cr",
        "Rs Crore": "inr_cr",
        "US $ Million": "usd_mn",
        "US$ Million": "usd_mn",
        "USD Million": "usd_mn",
        "Volume (Metric Tonnes)": "tonnes",
        # SDRs is an IMF unit; mapped to usd_mn (closest available). Promote
        # to sdr_mn dim_unit row when this lands in prod.
        "SDRs Million": "usd_mn",
    }
    base = target["imdr_prefix"]

    # Locate the period-label row: first row whose data cols contain a year/FY.
    header_row_idx = None
    for i, r in enumerate(rows):
        if len(r) >= data_first_col and any(
            re.match(r"^\d{4}(-\d{2,4})?$", (c or "").strip())
            for c in r[data_first_col:]
        ):
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], []

    col_period = _col_periods_from_header_block(
        rows, header_row_idx, data_first_col, n_header_rows=3,
    )
    if not col_period:
        return [], []

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()
    current_item = None

    for r in rows[header_row_idx + 3:]:
        if not r or len(r) < data_first_col + 1:
            continue
        item_cell = (r[1] or "").strip().replace("\xa0", " ") if len(r) > 1 else ""
        unit_cell = (r[2] or "").strip() if len(r) > 2 else ""
        if any(item_cell.lower().startswith(p)
                for p in ("source", "note", "p:", "p :")):
            break

        if item_cell:
            current_item = item_cell
        if current_item is None or not unit_cell:
            continue

        unit = unit_to_imdr.get(unit_cell)
        if unit is None:
            continue

        slug_raw = re.sub(r"^[\d.]+\s*", "", current_item)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug_raw).strip("_").upper()[:40] or "ROW"
        unit_slug = re.sub(r"[^A-Za-z0-9]+", "_", unit_cell).strip("_").upper()[:14]
        imdr_code = f"{base}.{slug}.{unit_slug}.IN"
        if imdr_code not in seen_codes:
            seen_codes.add(imdr_code)
            indicators.append(IndicatorRow(
                imdr_code=imdr_code, vendor_name="RBI",
                source_code=f"bulletin/{target['name']}/{slug}/{unit_slug}",
                display_name=(
                    f"{target['description']} — {current_item} ({unit_cell})"
                )[:255],
                unit=unit, frequency=target["frequency"], country_iso="IN",
                category=target["category"],
                is_seasonally_adjusted=False, bbg_ticker=None,
            ))
        for ci, period in col_period.items():
            if ci >= len(r):
                continue
            cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
            cell = cell.replace(",", "").replace("–", "").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=period, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_bop(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T40 BoP — Credit/Debit/Net × 2 quarter columns.

    Layout:
      R3 | Item | Q1-label   | …    | …    | Q2-label   | …    | …
      R4 |      | Credit     | Debit | Net  | Credit     | Debit | Net
    """
    base = target["imdr_prefix"]
    header_row_idx = None
    for i, r in enumerate(rows):
        if len(r) > 1 and (r[1] or "").strip() == "Item":
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], []
    header_row = rows[header_row_idx]
    sub_header = rows[header_row_idx + 1] if header_row_idx + 1 < len(rows) else []

    qtr_starts: dict[int, datetime.date] = {}
    last_dt: datetime.date | None = None
    for j in range(2, len(header_row)):
        cell = (header_row[j] or "").strip().replace("\xa0", " ")
        if cell:
            cell = re.sub(r"\s*\([PR]\)\s*$", "", cell).strip()
            m = re.match(r"^([A-Z][a-z]{2})-([A-Z][a-z]{2})\s+(\d{4})$", cell)
            if m:
                end_mo, year = m.group(2), m.group(3)
                try:
                    last_dt = datetime.datetime.strptime(
                        f"{end_mo} {year}", "%b %Y").date()
                except ValueError:
                    last_dt = None
        if last_dt is not None:
            qtr_starts[j] = last_dt
    col_kind: dict[int, str] = {}
    for j in range(2, len(sub_header)):
        kind = (sub_header[j] or "").strip()
        if kind in ("Credit", "Debit", "Net"):
            col_kind[j] = kind.upper()

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for r in rows[header_row_idx + 3:]:
        if not r or len(r) < 3:
            continue
        item = (r[1] or "").strip().replace("\xa0", " ") if len(r) > 1 else ""
        if not item or item.lower().startswith(("source", "note", "p:", "p :")):
            continue
        slug_raw = re.sub(r"^[\d.]+\s*", "", item)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug_raw).strip("_").upper()[:40] or "ROW"

        for j, kind in col_kind.items():
            if j >= len(r):
                continue
            cell = (r[j] if isinstance(r[j], str) else str(r[j])).strip()
            cell = cell.replace(",", "").replace("–", "-").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            period = qtr_starts.get(j)
            if period is None:
                continue
            imdr_code = f"{base}.{slug}.{kind}.IN"
            if imdr_code not in seen_codes:
                seen_codes.add(imdr_code)
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="RBI",
                    source_code=f"bulletin/{target['name']}/{slug}/{kind}",
                    display_name=(
                        f"{target['description']} — {item} — {kind}"
                    )[:255],
                    unit="usd_mn", frequency=target["frequency"],
                    country_iso="IN", category=target["category"],
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=period, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


# ---------------------------------------------------------------------------
# NEW parsers
# ---------------------------------------------------------------------------

def parse_nri_deposits_34(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T34 Non-Resident Deposits — dual-block Outstanding | Flows.

    Layout (May-2026):
      R2: | Scheme | Outstanding (cols 2-5) | Flows (cols 6-7)
      R3: |        | 2024-25  | 2025 |  2026 |       | 2024-25 | 2025-26
      R4: |        |          | Mar. | Feb.  | Mar.(P)| Apr.-Mar.| Apr.-Mar.(P)
      R5: |        | 1        | 2    | 3     | 4      | 5        | 6
      Data rows: 1. NRI Deposits / 1.1 FCNR(B) / 1.2 NR(E)RA / 1.3 NRO

    OUTSTANDING cols → month-end obs_date.
    FLOW cols → FY-end (31 March) obs_date with measure suffix FLOW to
    avoid collision with OUTSTANDING.

    Known check: FCNR(B) Mar-2026(P) outstanding = 33756, FY25-26 flow = 946.
    """
    base = target["imdr_prefix"]

    # Find the block-header row: contains both "Outstanding" and "Flow".
    block_row_idx = None
    for i, r in enumerate(rows):
        joined = " | ".join(c for c in r).lower()
        if "outstanding" in joined and "flow" in joined:
            block_row_idx = i
            break
    if block_row_idx is None:
        return [], []

    # Identify the column where "Flows" starts (to split OUTSTANDING vs FLOW).
    flows_start_col: int | None = None
    for ci, c in enumerate(rows[block_row_idx]):
        if "flow" in (c or "").lower():
            flows_start_col = ci
            break

    # Parse period labels from the next 4 header rows.
    # OUTSTANDING cols → month-end obs_date; FLOW cols → FY-end (Mar 31) date.
    #
    # T34 has a tricky layout: year "2025" and "2026" are in one header row,
    # and month tokens "Mar.", "Feb.", "Mar. (P)" are in the row below — BUT
    # the year label and the month token are in DIFFERENT columns (year at
    # col N, month at col N or col N+1). We handle this by pre-scanning every
    # header row left-to-right to build col_best_year[j] = the most recent
    # year seen in any row up to column j.
    col_period: dict[int, datetime.date] = {}
    col_is_flow: dict[int, bool] = {}

    hdr = rows[block_row_idx: block_row_idx + 4]
    max_cols = max((len(r) for r in hdr), default=0)

    # Build col_best_year: for each col j, the year visible at or before col j
    # in any header row. Scan left-to-right across all header rows.
    col_best_year: dict[int, str] = {}
    running_year: str | None = None
    for j in range(2, max_cols):
        for r in hdr:
            cell = (r[j] if j < len(r) else "").strip()
            if re.match(r"^\d{4}$", cell):
                running_year = cell
        if running_year:
            col_best_year[j] = running_year

    for j in range(2, max_cols):
        last_dt: datetime.date | None = None
        for r in hdr:
            cell = (r[j] if j < len(r) else "").strip()
            if not cell:
                continue
            if "outstanding" in cell.lower() or "flow" in cell.lower():
                continue
            # FY label "2024-25" / "2025-26" → FY-end March 31 of end-year
            fy_m = re.match(r"^(\d{4})-(\d{2,4})$", cell)
            if fy_m:
                year2 = fy_m.group(2)
                full_year = (int(fy_m.group(1)) + 1 if len(year2) == 2
                             else int(year2))
                last_dt = datetime.date(full_year, 3, 31)
                continue
            # Year-only token → already captured in col_best_year
            if re.match(r"^\d{4}$", cell):
                continue
            d = _parse_date(cell)
            if d is not None:
                last_dt = d
                continue
            # Month token ("Mar.", "Feb.", "Mar. (P)") — combine with best year
            yr = col_best_year.get(j)
            if yr:
                t = re.sub(r"\s*\([PR]\)\s*$", "", cell).strip().rstrip(".").strip()
                for fmt in ("%b %d %Y", "%b. %d %Y", "%b %Y", "%b. %Y"):
                    try:
                        last_dt = datetime.datetime.strptime(
                            f"{t} {yr}", fmt).date()
                        break
                    except ValueError:
                        continue
        if last_dt is not None:
            col_period[j] = last_dt
            col_is_flow[j] = (
                flows_start_col is not None and j >= flows_start_col
            )

    if not col_period:
        return [], []

    # Scheme name → slug map
    _SCHEME_MAP = {
        "fcnr": "FCNRB",
        "nr(e)ra": "NRERA",
        "nre": "NRERA",
        "nro": "NRO",
        "nri deposit": "NRI_TOTAL",
    }

    def _scheme_slug(label: str) -> str | None:
        low = re.sub(r"^[\d.\s]+", "", label).lower()
        for key, slug in _SCHEME_MAP.items():
            if key in low:
                return slug
        return None

    # Pre-register all indicators (4 schemes × 2 measures = 8)
    schemes = ["NRI_TOTAL", "FCNRB", "NRERA", "NRO"]
    indicators: list[IndicatorRow] = []
    for scheme in schemes:
        for measure in ("OUTSTANDING", "FLOW"):
            indicators.append(IndicatorRow(
                imdr_code=f"{base}.{scheme}.{measure}.IN",
                vendor_name="RBI",
                source_code=f"bulletin/T34/{scheme}/{measure}",
                display_name=(
                    f"RBI Bulletin T34 NRI Deposits — {scheme} {measure} (USD Mn)"
                )[:255],
                unit="usd_mn", frequency="MONTHLY", country_iso="IN",
                category="bop", is_seasonally_adjusted=False, bbg_ticker=None,
            ))

    observations: list[ObservationRow] = []
    for r in rows:
        if not r or len(r) < 3:
            continue
        label_raw = (r[1] or "").strip() if len(r) > 1 else ""
        scheme = _scheme_slug(label_raw)
        if scheme is None:
            continue
        for ci, period in col_period.items():
            if ci >= len(r):
                continue
            cell = (r[ci] or "").replace(",", "").strip()
            if not re.match(r"^-?\d+(\.\d+)?$", cell):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            measure = "FLOW" if col_is_flow.get(ci) else "OUTSTANDING"
            # OUTSTANDING is a point-in-time stock. The FY-label column
            # ("2024-25") resolves to Mar-31 and the explicit "Mar." month
            # column resolves to Mar-01 — same period, two dates, identical
            # value. Normalising OUTSTANDING to month-start collapses BOTH to
            # Mar-01 so the run_fetch dedup (keyed on imdr_code/obs_date/
            # vintage) keeps a single row. FLOW keeps its FY-end (31 Mar) date
            # — distinct FYs never collide.
            obs_date = period if measure == "FLOW" else period.replace(day=1)
            observations.append(ObservationRow(
                imdr_code=f"{base}.{scheme}.{measure}.IN",
                obs_date=obs_date, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_cd_cp(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T28 Certificates of Deposit / T29 Commercial Paper.

    Layout (both tables identical shape, 7 rows):
      R1: header row with "Item" in col 1
      R2: month tokens, no FY prefix (carry year from R1 FY cells)
      R3: column numbers
      Data rows:
        "1. Amount Outstanding (₹ Crore)"  → numeric, unit inr_cr
        "1.1 Issued during … (₹ Crore)"   → numeric, unit inr_cr
        "2. Rate of Interest (per cent)"   → range string "5.25-7.56"
                                             → emit RANGE_LO + RANGE_HI (pct)

    Rate rows reuse the range-split logic from parse_call_money_27.
    """
    base = target["imdr_prefix"]

    # Find header row (col 1 == "Item")
    header_row_idx = None
    for i, r in enumerate(rows):
        if len(r) >= 2 and (r[1] or "").strip() == "Item":
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], []

    col_period = _col_periods_from_header_block(
        rows, header_row_idx, data_first_col=2, n_header_rows=3,
    )
    if not col_period:
        return [], []

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    # data_start: first row after the column-number row (last header row)
    data_start = header_row_idx + 1
    for di in range(header_row_idx + 1, min(header_row_idx + 6, len(rows))):
        r = rows[di]
        if all(re.match(r"^\d+$", (c or "").strip()) for c in r[2:] if (c or "").strip()):
            data_start = di + 1
            break

    for r in rows[data_start:]:
        if not r or len(r) < 3:
            continue
        label = (r[1] or "").strip()
        if not label:
            continue
        if any(label.lower().startswith(p) for p in ("source", "note", "p:", "p :")):
            break

        is_rate_row = "rate" in label.lower() and "interest" in label.lower()
        slug_raw = re.sub(r"^[\d.]+\s*", "", label)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug_raw).strip("_").upper()[:50] or "ROW"
        # Strip unit in parens from slug — "(₹ Crore)" → already stripped by non-alnum
        unit = "pct" if is_rate_row else "inr_cr"

        if is_rate_row:
            # Range split: emit RANGE_LO + RANGE_HI
            for suffix in ("RANGE_LO", "RANGE_HI"):
                imdr_code = f"{base}.{slug}.{suffix}.IN"
                if imdr_code not in seen_codes:
                    seen_codes.add(imdr_code)
                    indicators.append(IndicatorRow(
                        imdr_code=imdr_code, vendor_name="RBI",
                        source_code=f"bulletin/{target['name']}/{slug}/{suffix}",
                        display_name=f"{target['description']} — {label} {suffix}"[:255],
                        unit="pct", frequency=target["frequency"], country_iso="IN",
                        category=target["category"],
                        is_seasonally_adjusted=False, bbg_ticker=None,
                    ))
            for ci, period in col_period.items():
                if ci >= len(r):
                    continue
                cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
                m = re.match(r"^\s*([\d.]+)\s*-\s*([\d.]+)\s*$", cell)
                if not m:
                    continue
                try:
                    lo, hi = float(m.group(1)), float(m.group(2))
                except ValueError:
                    continue
                for suffix, value in (("RANGE_LO", lo), ("RANGE_HI", hi)):
                    observations.append(ObservationRow(
                        imdr_code=f"{base}.{slug}.{suffix}.IN",
                        obs_date=period, vintage=0,
                        release_date=now, value=value, ingested_at=now,
                    ))
        else:
            imdr_code = f"{base}.{slug}.IN"
            if imdr_code not in seen_codes:
                seen_codes.add(imdr_code)
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="RBI",
                    source_code=f"bulletin/{target['name']}/{slug}",
                    display_name=f"{target['description']} — {label}"[:255],
                    unit=unit, frequency=target["frequency"], country_iso="IN",
                    category=target["category"],
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
            for ci, period in col_period.items():
                if ci >= len(r):
                    continue
                cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
                cell = cell.replace(",", "").replace("–", "").strip()
                if cell in ("", "-", "..", "NA", "N/A", "*"):
                    continue
                try:
                    value = float(cell)
                except ValueError:
                    continue
                observations.append(ObservationRow(
                    imdr_code=imdr_code, obs_date=period, vintage=0,
                    release_date=now, value=value, ingested_at=now,
                ))
    return indicators, observations


def parse_date_rows(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T3 Liquidity Operations by RBI (LAF) — DATE-PER-ROW layout.

    The header row is detected by finding the row whose col 1 == "Date".
    The row immediately after contains the series names (Repo, Reverse Repo,
    Variable Rate Repo, Variable Rate Reverse Repo, MSF, SDF, …).
    A second header row carries section group names (Liquidity Adjustment
    Facility, Standing Liquidity Facilities, OMO (Outright)).
    Series names from R3 (col 2+) are used as-is; "-" cells → skip.
    """
    base = target["imdr_prefix"]

    # Locate the "Date" header row.
    date_row_idx = None
    for i, r in enumerate(rows):
        if len(r) >= 2 and (r[1] or "").strip() == "Date":
            date_row_idx = i
            break
    if date_row_idx is None:
        return [], []

    # The next row has the series names.
    series_row = rows[date_row_idx + 1] if date_row_idx + 1 < len(rows) else []
    # Skip the column-number row (all numeric) to find data start.
    data_start = date_row_idx + 2
    for di in range(date_row_idx + 2, min(date_row_idx + 5, len(rows))):
        r = rows[di]
        if all(re.match(r"^\d+$", (c or "").strip()) for c in r[2:] if (c or "").strip()):
            data_start = di + 1
            break

    # Build col → series_name map from series_row (col 2+).
    col_series: dict[int, str] = {}
    for ci, v in enumerate(series_row):
        if ci < 2:
            continue
        name = (v or "").strip()
        if name:
            col_series[ci] = name

    if not col_series:
        return [], []

    # Also carry the group header from `date_row_idx` row (col 2+ spans).
    # We include the group prefix in the slug for disambiguation
    # (e.g., "Standing Liquidity Facilities" spans OMO columns).
    group_row = rows[date_row_idx]
    col_group: dict[int, str] = {}
    last_group = ""
    for ci, v in enumerate(group_row):
        if ci < 2:
            continue
        cell = (v or "").strip()
        if cell:
            last_group = cell
        col_group[ci] = last_group

    indicators: list[IndicatorRow] = []
    seen_codes: set[str] = set()
    for ci, series_name in col_series.items():
        group = col_group.get(ci, "")
        if group and group.lower() not in ("date", ""):
            full_name = f"{group} — {series_name}"
        else:
            full_name = series_name
        slug = re.sub(r"[^A-Za-z0-9]+", "_", full_name).strip("_").upper()[:50] or "COL"
        imdr_code = f"{base}.{slug}.IN"
        if imdr_code not in seen_codes:
            seen_codes.add(imdr_code)
            indicators.append(IndicatorRow(
                imdr_code=imdr_code, vendor_name="RBI",
                source_code=f"bulletin/{target['name']}/{slug}",
                display_name=f"{target['description']} — {full_name}"[:255],
                unit="inr_cr", frequency="DAILY", country_iso="IN",
                category=target["category"],
                is_seasonally_adjusted=False, bbg_ticker=None,
            ))

    observations: list[ObservationRow] = []
    for r in rows[data_start:]:
        if not r or len(r) < 3:
            continue
        date_cell = (r[1] or "").strip() if len(r) > 1 else ""
        if not date_cell:
            continue
        if any(date_cell.lower().startswith(p)
               for p in ("source", "note", "total")):
            break
        d = _parse_date(date_cell)
        if d is None:
            continue
        for ci, series_name in col_series.items():
            if ci >= len(r):
                continue
            cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
            cell = cell.replace(",", "").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            # Negative injection sign (e.g. "-179") is valid
            try:
                value = float(cell)
            except ValueError:
                continue
            group = col_group.get(ci, "")
            if group and group.lower() not in ("date", ""):
                full_name = f"{group} — {series_name}"
            else:
                full_name = series_name
            slug = re.sub(r"[^A-Za-z0-9]+", "_", full_name).strip("_").upper()[:50]
            observations.append(ObservationRow(
                imdr_code=f"{base}.{slug}.IN",
                obs_date=d, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_iip_assets_liab(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T44 International Investment Position — paired Assets/Liabilities columns.

    Layout:
      R3: | | | 2024-25 |    | 2024 |    | 2025 |
      R4: | | |         |    | Dec. |    | Sep. |    | Dec.
      R5: | | | Assets  | Liabilities | Assets | Liabilities | ...
      R6: col numbers 1..8
      Data rows start R7.

    Each period has two columns (Assets, Liabilities). Emit
    <slug>.ASSETS and <slug>.LIABILITIES indicators per item.
    Rows with only Assets (e.g. "4. Reserves") or only Liabilities
    (e.g. "3.5 SDR net incurrence") are handled naturally — the absent
    column is empty/blank and gets skipped.
    """
    base = target["imdr_prefix"]

    # Find the "Item" label row (R2 in May-2026).
    item_row_idx = None
    for i, r in enumerate(rows):
        if len(r) >= 2 and (r[1] or "").strip() == "Item":
            item_row_idx = i
            break
    if item_row_idx is None:
        return [], []

    # R3 = FY/year labels, R4 = month tokens, R5 = Assets/Liabilities labels.
    # Resolve period for each column pair from the year+month header rows.
    # data_first_col = 2
    data_first_col = 2

    # Build year+month date per column from R3+R4 (2 rows above the A/L row).
    # Then read A/L from R5.
    col_period = _col_periods_from_header_block(
        rows, item_row_idx, data_first_col, n_header_rows=4,
    )

    # Read the Assets/Liabilities sub-header from the row that contains those tokens.
    al_row_idx = None
    for i in range(item_row_idx + 1, min(item_row_idx + 6, len(rows))):
        r = rows[i]
        joined = " ".join(c for c in r).strip()
        if "Assets" in joined and "Liabilities" in joined:
            al_row_idx = i
            break
    if al_row_idx is None:
        return [], []

    al_row = rows[al_row_idx]
    col_kind: dict[int, str] = {}  # col_idx → "ASSETS" or "LIABILITIES"
    for ci, v in enumerate(al_row):
        cell = (v or "").strip()
        if cell == "Assets":
            col_kind[ci] = "ASSETS"
        elif cell == "Liabilities":
            col_kind[ci] = "LIABILITIES"

    # For each col in col_kind, pair it with the period from col_period.
    # col_period resolves dates from the year/month header rows. The period
    # columns in T44 interleave: (A24-25, L24-25, A_Dec24, L_Dec24, ...);
    # both A and L for the same quarter share the same date.
    # We propagate the most recent resolved date leftward to fill gaps.
    paired: dict[int, tuple[datetime.date, str]] = {}
    for ci in range(data_first_col, max(col_kind.keys(), default=0) + 2):
        if ci in col_kind and ci in col_period:
            paired[ci] = (col_period[ci], col_kind[ci])
        elif ci in col_kind:
            # No direct period: look left for nearest period
            for offset in range(1, 5):
                prev = ci - offset
                if prev in col_period:
                    paired[ci] = (col_period[prev], col_kind[ci])
                    break

    if not paired:
        return [], []

    # data_start: skip past col-number row
    data_start = al_row_idx + 1
    for di in range(al_row_idx + 1, min(al_row_idx + 4, len(rows))):
        r = rows[di]
        if all(re.match(r"^\d+$", (c or "").strip()) for c in r[2:] if (c or "").strip()):
            data_start = di + 1
            break

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for r in rows[data_start:]:
        if not r or len(r) < data_first_col + 1:
            continue
        label = (r[1] or "").strip().replace("\xa0", " ")
        if not label:
            continue
        if any(label.lower().startswith(p) for p in ("source", "note", "note:", "explanatory")):
            break

        slug_raw = re.sub(r"^[\d.]+\s*", "", label)
        slug = re.sub(r"[^A-Za-z0-9]+", "_", slug_raw).strip("_").upper()[:40] or "ROW"

        for ci, (period, kind) in paired.items():
            if ci >= len(r):
                continue
            cell = (r[ci] if isinstance(r[ci], str) else str(r[ci])).strip()
            cell = cell.replace(",", "").replace("–", "").strip()
            if cell in ("", "-", "..", "NA", "N/A", "*"):
                continue
            try:
                value = float(cell)
            except ValueError:
                continue
            imdr_code = f"{base}.{slug}.{kind}.IN"
            if imdr_code not in seen_codes:
                seen_codes.add(imdr_code)
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="RBI",
                    source_code=f"bulletin/{target['name']}/{slug}/{kind}",
                    display_name=(
                        f"{target['description']} — {label} — {kind}"
                    )[:255],
                    unit="usd_mn", frequency=target["frequency"],
                    country_iso="IN", category=target["category"],
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=period, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def parse_tbill_auctions_26(
    rows: list[list[str]], target: dict, now: datetime.datetime,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    """T26 Auctions of Treasury Bills — per-tenor per-auction-date rows.

    Layout:
      R2: Date of Auction | Notified Amount | Bids Received (Number/CompFV/NonCompFV) |
          Bids Accepted (Number/CompFV/NonCompFV) | Total Issue | Cut-off Price |
          Implicit Yield at Cut-off Price (per cent)
      R3: sub-headers (Number, Total Face Value)
      R4: competitive/non-competitive split
      Tenor sections: "91-day Treasury Bills", "182-day Treasury Bills",
                      "364-day Treasury Bills" (value-less rows)
      FY subheader rows: "2025-26" (value-less)
      Data rows: auction dates like "Feb. 25", "Mar. 4" (partial date — no year;
                 must combine with FY subheader year context).

    Extracted per auction date + tenor:
      - Notified Amount (col 2, inr_cr)
      - Bids Accepted Total Face Value = Competitive + Non-competitive (cols 7+8)
      - Implicit Yield at Cut-off Price (col 11, pct)

    Cut-off Price (col 10) is not extracted (redundant — yield is the desk signal).

    TODO: The "Total Issue (6+7)" column (col 9) overlaps with Bids Accepted
    face value; confirm col numbering if the Bulletin layout changes in future months.
    """
    base = target["imdr_prefix"]

    # Column layout (0-indexed from _parse_sheet output):
    # col 1: Date of Auction (or section header)
    # col 2: Notified Amount
    # col 3: Bids Received Number
    # col 4: Bids Received FV Competitive
    # col 5: Bids Received FV Non-Competitive
    # col 6: Bids Accepted Number
    # col 7: Bids Accepted FV Competitive
    # col 8: Bids Accepted FV Non-Competitive
    # col 9: Total Issue (6+7)
    # col 10: Cut-off Price
    # col 11: Implicit Yield at Cut-off Price (%)
    COL_NOTIFIED = 2
    COL_ACCEPTED_COMP = 7
    COL_ACCEPTED_NONCOMP = 8
    COL_YIELD = 11

    # Pre-register indicators per tenor (3 tenors × 3 measures)
    TENORS = ["91D", "182D", "364D"]
    _TENOR_MAP = {
        "91-day": "91D",
        "182-day": "182D",
        "364-day": "364D",
    }
    indicators: list[IndicatorRow] = []
    seen_codes: set[str] = set()

    def _register(tenor: str, measure: str, unit: str, desc_suffix: str) -> str:
        imdr_code = f"{base}.{tenor}.{measure}.IN"
        if imdr_code not in seen_codes:
            seen_codes.add(imdr_code)
            indicators.append(IndicatorRow(
                imdr_code=imdr_code, vendor_name="RBI",
                source_code=f"bulletin/{target['name']}/{tenor}/{measure}",
                display_name=(
                    f"{target['description']} — {tenor} {desc_suffix}"
                )[:255],
                unit=unit, frequency=target["frequency"],
                country_iso="IN", category=target["category"],
                is_seasonally_adjusted=False, bbg_ticker=None,
            ))
        return imdr_code

    observations: list[ObservationRow] = []
    current_tenor: str | None = None
    current_fy_end_year: int | None = None  # the ending year of the FY (e.g. 2026 for 2025-26)

    # Find data_start (after header block)
    data_start = 0
    for i, r in enumerate(rows):
        if any(re.match(r"^-day", (c or "").strip().lower()) or
               re.match(r"^\d+-day", (c or "").strip().lower())
               for c in r):
            data_start = i
            break
        # Also look for the tenor section row pattern
        if len(r) >= 2:
            label = (r[1] or "").strip().lower()
            if any(k in label for k in ("91-day", "182-day", "364-day")):
                data_start = i
                break

    for i in range(data_start, len(rows)):
        r = rows[i]
        if not r:
            continue
        label = (r[1] or "").strip() if len(r) > 1 else ""
        if not label:
            continue
        if any(label.lower().startswith(p) for p in ("source", "note")):
            break

        # Section: tenor header ("91-day Treasury Bills")
        tenor_match = None
        for key, slug in _TENOR_MAP.items():
            if key in label.lower():
                tenor_match = slug
                break
        if tenor_match:
            current_tenor = tenor_match
            continue

        # FY subheader ("2025-26")
        fy_m = re.match(r"^(\d{4})-(\d{2,4})$", label)
        if fy_m:
            year2 = fy_m.group(2)
            current_fy_end_year = (
                int(fy_m.group(1)) + 1 if len(year2) == 2 else int(year2)
            )
            continue

        if current_tenor is None or current_fy_end_year is None:
            continue

        # Resolve auction date. Label is "Feb. 25" / "Mar. 4" (month + day);
        # the year is DERIVED from the FY subheader, not a blind "try both
        # years and take the first that parses". The Indian FY runs Apr
        # (start year) → Mar (end year), so Jan–Mar auctions fall in
        # current_fy_end_year and Apr–Dec auctions fall in the prior calendar
        # year (current_fy_end_year - 1). The old "first parseable year" logic
        # was wrong — an "Apr." auction in FY2025-26 resolved to 2026 (should
        # be 2025).
        date_str_raw = re.sub(r"\s*\([PR]\)\s*$", "", label).strip()
        md: datetime.date | None = None
        for fmt in ("%b. %d %Y", "%b %d %Y", "%B %d %Y"):
            try:
                md = datetime.datetime.strptime(
                    f"{date_str_raw} {current_fy_end_year}", fmt).date()
                break
            except ValueError:
                continue
        if md is None:
            continue
        yr = current_fy_end_year if md.month <= 3 else current_fy_end_year - 1
        try:
            auction_date = md.replace(year=yr)
        except ValueError:
            auction_date = md  # Feb-29 in a non-leap target year — rare; keep parsed

        def _cell(col: int) -> str:
            if col >= len(r):
                return ""
            return (r[col] if isinstance(r[col], str) else str(r[col])).strip().replace(",", "")

        def _float(col: int) -> float | None:
            c = _cell(col)
            if c in ("", "-", "..", "NA", "N/A", "*", "0"):
                return None if c in ("", "-", "..", "NA", "N/A", "*") else 0.0
            try:
                return float(c)
            except ValueError:
                return None

        notified = _float(COL_NOTIFIED)
        accepted_comp = _float(COL_ACCEPTED_COMP)
        accepted_noncomp = _float(COL_ACCEPTED_NONCOMP)
        yield_val = _float(COL_YIELD)

        accepted_total: float | None = None
        if accepted_comp is not None or accepted_noncomp is not None:
            accepted_total = (accepted_comp or 0.0) + (accepted_noncomp or 0.0)

        for measure, value, unit, desc in (
            ("NOTIFIED_AMT", notified, "inr_cr", "Notified Amount"),
            ("ACCEPTED_FV", accepted_total, "inr_cr", "Bids Accepted Face Value"),
            ("CUTOFF_YIELD", yield_val, "pct", "Implicit Yield at Cut-off"),
        ):
            if value is None:
                continue
            imdr_code = _register(current_tenor, measure, unit, desc)
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=auction_date, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))

    return indicators, observations


# ---------------------------------------------------------------------------
# Parser registry
# ---------------------------------------------------------------------------

_PARSERS = {
    "parse_call_money_27": parse_call_money_27,
    "parse_cpi_combined_19c": parse_cpi_combined_19c,
    "parse_wide_table": parse_wide_table,
    "parse_dual_unit": parse_dual_unit,
    "parse_bop": parse_bop,
    "parse_nri_deposits_34": parse_nri_deposits_34,
    "parse_cd_cp": parse_cd_cp,
    "parse_date_rows": parse_date_rows,
    "parse_iip_assets_liab": parse_iip_assets_liab,
    "parse_tbill_auctions_26": parse_tbill_auctions_26,
}


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def run_fetch(
    since: str | None,
    until: str | None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    since_dt = datetime.date.fromisoformat(since) if since else None
    until_dt = datetime.date.fromisoformat(until) if until else None
    now = datetime.datetime.now(UTC)

    # TSPD blocks headless — always run headed. See module docstring.
    saved = _download_via_headed(PRIORITY_TARGETS, headless=False)
    name_to_path = {p.stem: p for p in saved}

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen_codes: set[str] = set()

    for t in PRIORITY_TARGETS:
        path = name_to_path.get(t["name"])
        if path is None or not path.exists():
            print(f"  skip {t['name']} (no XLSX)")
            continue
        rows = _parse_sheet(path)
        parser = _PARSERS.get(t.get("parser"))
        if parser is None:
            print(f"  no parser registered for {t['name']}, skip")
            continue
        inds, obs = parser(rows, t, now)
        for i in inds:
            if i.imdr_code in seen_codes:
                continue
            seen_codes.add(i.imdr_code)
            indicators.append(i)
        for o in obs:
            if since_dt and o.obs_date < since_dt:
                continue
            if until_dt and o.obs_date > until_dt:
                continue
            observations.append(o)
        print(f"  {t['name']}: +{len(inds)} indicators / +{len(obs)} obs")

    # Dedup on the fact PK (imdr_code, obs_date, vintage). Several Bulletin
    # tables carry two columns that resolve to the same date — e.g. a full-FY
    # "2024-25" column and the "2024-25 Apr.-Mar." cumulative column both map
    # to 2024-04-01 (identical value, since India FY = Apr-Mar). Without this,
    # the loader's MERGE hits a PRIMARY KEY violation. Keep last-seen.
    deduped: dict[tuple[str, datetime.date, int], ObservationRow] = {}
    for o in observations:
        deduped[(o.imdr_code, o.obs_date, o.vintage)] = o
    n_dropped = len(observations) - len(deduped)
    if n_dropped:
        print(f"  deduped {n_dropped} same-PK observation(s)")
    return indicators, list(deduped.values())


def main() -> int:
    return run_main(vendor="rbi", topic="bulletin",
                    fetch_fn=run_fetch,
                    description=__doc__.splitlines()[0] if __doc__ else "",
                    country_code="IN")


if __name__ == "__main__":
    sys.exit(main())
