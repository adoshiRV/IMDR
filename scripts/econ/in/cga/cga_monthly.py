"""CGA monthly fiscal accounts — Centre receipts / expenditure / deficits.

Source: Controller General of Accounts dashboard at
``cga.nic.in/MonthDashboardReport/Published/list.aspx``. The DAMA
dashboard XLSM ships the FULL monthly history from FY 2014-15. Sheet
``actual`` carries month-by-month line items; ~30 series × ~143 months
≈ 4,200 obs.

Cell mapping (see docs/admin/econ/india/in_coverage_plan.md):
  4.1 Fiscal stance — all fields
"""
from __future__ import annotations

import datetime
import io
import re
import urllib.parse
import warnings

import httpx
from openpyxl import load_workbook

from imdr.domains.econ.schema import IndicatorRow, ObservationRow
from scripts.econ._runner import run_main

warnings.filterwarnings("ignore", category=UserWarning)

UTC = datetime.timezone.utc

_DASHBOARD_LIST = "https://cga.nic.in/MonthDashboardReport/Published/list.aspx"
_UA = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 Chrome/120 Safari/537.36"),
    "Referer": "https://cga.nic.in/",
}

_COLUMNS: dict[int, tuple[str, str]] = {
    2:  ("RECEIPT.CORPTAX",      "Corporation Tax — Centre actual receipts (CGA, INR crore)"),
    3:  ("RECEIPT.INCTAX",       "Income Tax — Centre actual receipts (CGA, INR crore)"),
    4:  ("RECEIPT.STT",          "Securities Transaction Tax — Centre actual receipts (CGA, INR crore)"),
    5:  ("RECEIPT.CGST",         "CGST — Centre actual receipts (CGA, INR crore)"),
    6:  ("RECEIPT.IGST",         "IGST — Centre actual receipts (CGA, INR crore)"),
    7:  ("RECEIPT.UTGST",        "UTGST — Centre actual receipts (CGA, INR crore)"),
    8:  ("RECEIPT.GST_COMPCESS", "GST Compensation Cess — Centre actual receipts (CGA, INR crore)"),
    9:  ("RECEIPT.CUSTOMS",      "Customs — Centre actual receipts (CGA, INR crore)"),
    10: ("RECEIPT.UNION_EXCISE", "Union Excise — Centre actual receipts (CGA, INR crore)"),
    11: ("RECEIPT.SERVICE_TAX",  "Service Tax — Centre actual receipts (CGA, INR crore — legacy)"),
    12: ("RECEIPT.OTHER_TAX",    "Other Taxes — Centre actual receipts (CGA, INR crore)"),
    13: ("RECEIPT.NCCF_SURCHG",  "NCCF Surcharge — Centre actual receipts (CGA, INR crore)"),
    14: ("EXPEND.DEVOLUTION",    "Devolution to States — Centre transfer (CGA, INR crore)"),
    15: ("RECEIPT.INTEREST_RX",  "Interest receipts — Centre non-tax (CGA, INR crore)"),
    16: ("RECEIPT.DIVIDENDS",    "Dividends and Profits — Centre non-tax (CGA, INR crore)"),
    17: ("RECEIPT.OTHER_NONTAX", "Other Non-Tax Receipts — Centre (CGA, INR crore)"),
    18: ("RECEIPT.LOAN_RECOVERY","Recovery of Loans & Advances — Centre (CGA, INR crore)"),
    19: ("RECEIPT.DISINVEST",    "Disinvestment Receipts — Centre (CGA, INR crore)"),
    20: ("EXPEND.REVENUE",       "Revenue Expenditure — Centre (CGA, INR crore)"),
    21: ("EXPEND.INTEREST_PMT",  "Interest Payments — Centre (CGA, INR crore)"),
    22: ("EXPEND.DEFENCE",       "Defence Services — Centre (CGA, INR crore)"),
    23: ("EXPEND.GRANTS_STATES", "Grants in Aid to States & UTs — Centre (CGA, INR crore)"),
    24: ("EXPEND.PENSIONS",      "Pensions — Centre (CGA, INR crore)"),
    25: ("EXPEND.MAJOR_SUBSIDY", "Major Subsidies — Centre (CGA, INR crore)"),
    26: ("EXPEND.GRANTS_CAP",    "Grants for Creation of Capital Assets — Centre (CGA, INR crore)"),
    27: ("EXPEND.CAPITAL",       "Capital Expenditure — Centre (CGA, INR crore)"),
    30: ("DEFICIT.REVENUE",      "Revenue Deficit — Centre (CGA, INR crore)"),
    31: ("DEFICIT.EFF_REVENUE",  "Effective Revenue Deficit — Centre (CGA, INR crore)"),
    32: ("DEFICIT.FISCAL",       "Fiscal Deficit — Centre (CGA, INR crore)"),
    33: ("DEFICIT.PRIMARY",      "Primary Deficit — Centre (CGA, INR crore)"),
}

_MONTH_RE = re.compile(r"^([A-Za-z]{3})-(\d{2})$")
# Bucketed under "other" until a dedicated "fiscal" code is added to
# econ.dim_indicator_category + VALID_CATEGORIES in
# src/imdr/domains/econ/schema.py. Tracked in
# docs/admin/econ/india/in_coverage_plan.md.
_FISCAL_CATEGORY = "other"


def _fetch_latest(client: httpx.Client) -> tuple[str, bytes]:
    r = client.get(_DASHBOARD_LIST, headers=_UA, timeout=30)
    r.raise_for_status()
    links = re.findall(r'href=["\']([^"\']*?\.xlsm)["\']', r.text, re.I)
    if not links:
        raise RuntimeError("no .xlsm link on CGA dashboard list page")
    rel = next((l for l in links if "dama dashboard" in l.lower()), links[0])
    url = ("https://cga.nic.in" + urllib.parse.quote(rel)) if rel.startswith("/") else rel
    rf = client.get(url, headers=_UA, timeout=60)
    rf.raise_for_status()
    return url, rf.content


def _parse_month(s: str) -> datetime.date | None:
    m = _MONTH_RE.match(s.strip())
    if not m:
        return None
    try:
        mm = datetime.datetime.strptime(m.group(1), "%b").month
    except ValueError:
        return None
    return datetime.date(2000 + int(m.group(2)), mm, 1)


def run_fetch(
    since: str | None,
    until: str | None,
) -> tuple[list[IndicatorRow], list[ObservationRow]]:
    since_dt = datetime.date.fromisoformat(since) if since else None
    until_dt = datetime.date.fromisoformat(until) if until else None
    now = datetime.datetime.now(UTC)

    with httpx.Client(timeout=60, follow_redirects=True) as c:
        url, blob = _fetch_latest(c)
        print(f"  {url}")
        print(f"  {len(blob)} bytes")

    wb = load_workbook(io.BytesIO(blob), data_only=True)
    ws = wb["actual"]
    print(f"  actual sheet: {ws.max_row}x{ws.max_column}")

    indicators: list[IndicatorRow] = []
    observations: list[ObservationRow] = []
    seen: set[str] = set()
    seen_obs: set[tuple[str, datetime.date]] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        obs_date = _parse_month(str(row[1]))
        if obs_date is None:
            continue
        if since_dt and obs_date < since_dt:
            continue
        if until_dt and obs_date > until_dt:
            continue
        for col_idx, (stem, display) in _COLUMNS.items():
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            try:
                value = float(v)
            except (ValueError, TypeError):
                continue
            imdr_code = f"INDIA.FISCAL.{stem}.IN"
            if imdr_code not in seen:
                indicators.append(IndicatorRow(
                    imdr_code=imdr_code, vendor_name="CGA",
                    source_code=f"CGA/MonthAccountDashboard/actual/col{col_idx}",
                    display_name=display, unit="inr_cr",
                    frequency="MONTHLY", country_iso="IN",
                    category=_FISCAL_CATEGORY,
                    is_seasonally_adjusted=False, bbg_ticker=None,
                ))
                seen.add(imdr_code)
            key = (imdr_code, obs_date)
            if key in seen_obs:
                continue
            seen_obs.add(key)
            observations.append(ObservationRow(
                imdr_code=imdr_code, obs_date=obs_date, vintage=0,
                release_date=now, value=value, ingested_at=now,
            ))
    return indicators, observations


def main() -> int:
    return run_main(vendor="cga", topic="monthly",
                    fetch_fn=run_fetch,
                    description=__doc__.splitlines()[0] if __doc__ else "",
                    country_code="IN")


if __name__ == "__main__":
    import sys
    sys.exit(main())
